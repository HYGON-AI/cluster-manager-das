#!/usr/bin/env python3
# Copyright (c) 2026 Hygon Information Technology Co., Ltd.
# SPDX-License-Identifier: Apache-2.0
"""
日志分析工具
用于分析分布式系统中的时间戳日志
"""

import re
import os
import shutil
import logging
from datetime import datetime, date, time, timedelta
from collections import defaultdict, OrderedDict
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any, Set
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import numpy as np
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import mmap
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.layout import Layout, ManualLayout


class LogAnalyzerConfig:
    """配置类，集中管理所有配置参数"""
    
    def __init__(self):
        self.node_length = 8
        self.ends_slogan = '_res.txt'
        self.left_time = '【'
        self.right_time = '】'
        self.default_search_dir = '/public/home/user/workspace/labs/hcu_megatron/Megatron-LM'
        self.slow_stage_threshold = 1.0  # 慢阶段阈值，单位：秒
        self.interval_time = 1.0
        self.filter_threshold_seconds = 5.0  # 过滤阈值，单位：秒
        self.num_nodes = 16 # 节点数量


class LogProcessor:
    """日志处理基类"""
    
    def __init__(self, config: LogAnalyzerConfig):
        self.config = config
        self.logger = self._setup_logging()
    
    def _setup_logging(self) -> logging.Logger:
        """设置日志记录"""
        logger = logging.getLogger(self.__class__.__name__)
        if not logger.handlers:
            handler = logging.StreamHandler()
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger
    
    def _ensure_directory_exists(self, directory: Path) -> None:
        """确保目录存在"""
        try:
            directory.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            self.logger.error(f"无法创建目录 {directory}: {e}")
            raise

class LogExtractor:
    """日志提取器，优化大文件处理能力"""
    
    def __init__(self, config):
        self.config = config
        self.batch_size = getattr(config, 'processing_batch_size', 10000)
        self.max_open_files = getattr(config, 'max_open_files', 1000)
        self.node_length = getattr(config, 'node_length', 8)
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def extract_lines_by_slogan(self, input_file: Path, start_slogan: str) -> Path:
        """流式处理大文件，避免内存爆炸"""
        if not input_file.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
        
        output_file = input_file.parent / 'all_nodes.txt'
        
        try:
            processed_count = 0
            with input_file.open('r', encoding='utf-8', errors='ignore') as infile, \
                 output_file.open('w', encoding='utf-8') as outfile:
                
                batch = []
                for line in infile:
                    if line.startswith(start_slogan):
                        processed_line = self._process_line(line, start_slogan)
                        if processed_line:
                            batch.append(processed_line)
                            processed_count += 1
                    
                    # 批量写入
                    if len(batch) >= self.batch_size:
                        outfile.write('\n'.join(batch) + '\n')
                        batch = []
                        if processed_count % (self.batch_size * 10) == 0:
                            self.logger.debug(f"已处理 {processed_count} 行")
                
                # 写入剩余批次
                if batch:
                    outfile.write('\n'.join(batch) + '\n')
            
            self.logger.info(f"成功提取 {processed_count} 行日志到: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"提取日志时出错: {e}")
            raise
    
    def split_logs_by_host(self, input_file: Path, scale_hint: Optional[int] = None) -> Path:
        """
        根据主机名分割日志文件，自动选择最适合的分割策略
        
        Args:
            input_file: 输入文件路径
            scale_hint: 规模提示，可以是预估的主机数量或文件大小(MB)，用于自动选择策略
        
        Returns:
            输出目录路径
        """
        if not input_file.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
        
        
        # 根据规模选择策略
        strategy = self._select_split_strategy(scale_hint)
        self.logger.info(f"选择分割策略: {strategy}, 规模提示: {scale_hint}")
        
        # 执行分割
        if strategy == "basic":
            return self._split_logs_basic(input_file)
        elif strategy == "streaming":
            return self._split_logs_streaming(input_file)
        else:
            raise ValueError(f"未知的分割策略: {strategy}")
    

    def _select_split_strategy(self, scale_hint: int) -> str:
        """根据规模提示选择分割策略"""
        if scale_hint < 1024:
            return "basic"  # 小规模：主机数<64或文件<64MB
        else:
            return "streaming"  # 中等规模：主机数<1024或文件<1GB

    
    def _split_logs_basic(self, input_file: Path) -> Path:
        """基础分割实现 - 适用于小规模数据"""
        out_dir = input_file.parent / 'logs'
        self._prepare_output_dir(out_dir)
        
        host_lines = defaultdict(list)
        
        try:
            with input_file.open('r', encoding='utf-8') as infile:
                for line in infile:
                    if 'host:' in line:
                        host = line.split('host:')[-1].strip()
                        host_lines[host].append(line.strip())
            
            # 为每个主机创建文件
            for host, lines in host_lines.items():
                output_file = out_dir / f"{host}.txt"
                with output_file.open('w', encoding='utf-8') as outfile:
                    for line in lines:
                        outfile.write(f"{line}\n")
            
            self.logger.info(f"使用基础策略成功分割日志到 {len(host_lines)} 个主机文件")
            return out_dir
            
        except Exception as e:
            self.logger.error(f"基础分割失败: {e}")
            raise
    
    def _split_logs_streaming(self, input_file: Path) -> Path:
        """流式分割实现 - 适用于中等规模数据"""
        out_dir = input_file.parent / 'logs'
        self._prepare_output_dir(out_dir)
        
        file_handles = {}
        processed_count = 0
        
        try:
            with input_file.open('r', encoding='utf-8', errors='ignore') as infile:
                for line in infile:
                    if 'host:' in line:
                        host = line.split('host:')[-1].strip()
                        
                        if host not in file_handles:
                            output_file = out_dir / f"{host}.txt"
                            file_handles[host] = output_file.open('w', encoding='utf-8')
                        
                        file_handles[host].write(f"{line.strip()}\n")
                        processed_count += 1
                    
                    if processed_count % self.batch_size == 0:
                        self._cleanup_file_handles(file_handles, self.max_open_files)
                        if processed_count % (self.batch_size * 10) == 0:
                            self.logger.debug(f"流式处理已处理 {processed_count} 行")
            
            self._close_all_file_handles(file_handles)
            self.logger.info(f"使用流式策略成功分割 {processed_count} 行日志到 {len(file_handles)} 个主机文件")
            return out_dir
            
        except Exception as e:
            self._close_all_file_handles(file_handles)
            self.logger.error(f"流式分割失败: {e}")
            raise

    def _process_line(self, line: str, start_slogan: str) -> Optional[str]:
        """处理单行日志"""
        start_index = line.find(start_slogan)
        host_index = line.find("host:")
        
        if host_index == -1:
            return None
        
        # 构建相关部分
        relevant_part = line[start_index:host_index + 5]  # 保留 "host:"
        host_data = line[host_index + 5:]
        
        # 处理主机数据
        if len(host_data) < self.node_length:
            relevant_part += host_data
        else:
            relevant_part += host_data[:self.node_length]
        
        return relevant_part.strip()
    
    def _prepare_output_dir(self, out_dir: Path):
        """准备输出目录"""
        if out_dir.exists():
            shutil.rmtree(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
    

    def _cleanup_file_handles(self, file_handles: Dict, max_handles: int):
        """清理文件句柄"""
        if len(file_handles) <= max_handles:
            return
        
        # 关闭最早打开的文件句柄
        hosts_to_close = list(file_handles.keys())[:len(file_handles) - max_handles // 2]
        for host in hosts_to_close:
            try:
                file_handles[host].close()
                del file_handles[host]
            except Exception as e:
                self.logger.warning(f"关闭文件句柄 {host} 时出错: {e}")
    
    def _close_all_file_handles(self, file_handles: Dict):
        """关闭所有文件句柄"""
        for host, handle in list(file_handles.items()):
            try:
                handle.close()
                del file_handles[host]
            except Exception as e:
                self.logger.warning(f"关闭文件句柄 {host} 时出错: {e}")






class LogParser(LogProcessor):
    """日志解析器，负责解析时间戳信息"""
    
    def parse_single_log(self, input_file: Path) -> Path:
        """解析单个日志文件的时间戳信息"""
        if not input_file.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
        
        node_name = input_file.stem
        output_file = input_file.parent / f"{input_file.stem}{self.config.ends_slogan}"
        
        start_times = defaultdict(list)
        date = None
        
        try:
            with input_file.open('r', encoding='utf-8') as infile:
                for line in infile:
                    date, timestamp = self._extract_timestamp(line)
                    if timestamp and date:
                        start_info = self._extract_start_info(line)
                        start_times[start_info].append(timestamp)
            
            self._write_analysis_result(output_file, node_name, start_times, date)
            self.logger.info(f"解析结果已保存到: {output_file}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"解析日志 {input_file} 时出错: {e}")
            raise
    
    def _extract_timestamp(self, line: str) -> Tuple[Optional[str], Optional[str]]:
        """从行中提取时间戳和日期"""
        parts = line.split(" ")
        if len(parts) < 3:
            return None, None
        
        try:
            timestamp = parts[2].split(self.config.right_time)[0]
            date = parts[1].split(self.config.left_time)[-1]
            return date, timestamp
        except (IndexError, ValueError):
            return None, None
    
    def _extract_start_info(self, line: str) -> str:
        """提取起始信息"""
        return line.split("host:")[0].split(self.config.right_time)[-1]
    
    def _write_analysis_result(self, output_file: Path, node_name: str, 
                             start_times: Dict[str, List[str]], date: Optional[str]) -> None:
        """写入分析结果"""
        with output_file.open('w', encoding='utf-8') as outfile:
            for start, times in start_times.items():
                if not date or not times:
                    continue
                
                min_time_str = f"{date} {min(times)}"
                max_time_str = f"{date} {max(times)}"
                
                try:
                    min_time = datetime.strptime(min_time_str, "%Y-%m-%d %H:%M:%S.%f")
                    max_time = datetime.strptime(max_time_str, "%Y-%m-%d %H:%M:%S.%f")
                    time_diff = max_time - min_time
                    
                    outfile.write(f"节点【{node_name}】 - {start}的所有时间戳：{times}\n")
                    outfile.write(f"{start} - 最早时间: {min_time}, 最晚时间: {max_time}，相差：{time_diff}\n")
                except ValueError as e:
                    self.logger.warning(f"时间格式解析错误: {e}")


class LogAnalyzer(LogProcessor):
    """日志分析器，负责综合分析多个日志文件"""
    
    def __init__(self, config: LogAnalyzerConfig):
        super().__init__(config)
        self.start_keys: List[str] = []
        self.gt_threshold_times: List[str] = []
        # self.simple_result_content: List[str] = []
        self._init_excel_formats()

    def _init_excel_formats(self):
        """初始化Excel格式"""
        self.red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def analyze_logs_directory(self, logs_dir: Path) -> Tuple[Path, Path]:
        """分析日志目录中的所有结果文件"""
        if not logs_dir.is_dir():
            raise NotADirectoryError(f"目录不存在: {logs_dir}")
        
        result_files = list(logs_dir.glob(f"*{self.config.ends_slogan}"))
        if not result_files:
            raise FileNotFoundError(f"{logs_dir} 中没有后缀为 {self.config.ends_slogan} 的日志文件")
        
        # 提取节点名称
        node_names = [f.stem.replace(self.config.ends_slogan.replace('.txt', ''), '') 
                     for f in result_files]
        
        # 并行处理文件解析
        all_data, all_times = self._parse_all_files(result_files, node_names)
        
        if not all_data:
            raise RuntimeError("所有文件解析失败，没有有效数据")
        
        # 设置start_keys
        first_node = next(iter(all_data.keys()))
        self.start_keys = list(all_data[first_node].keys())

        # 生成结果文件
        result_paths = self._generate_output_files(logs_dir, all_data, all_times)
        
        self.logger.info(f"分析完成，结果保存在: {result_paths['result_txt']}、{result_paths['result_excel']}")
        return result_paths['result_txt'], result_paths['result_excel']

    def _parse_all_files(self, result_files: List[Path], node_names: List[str]) -> Tuple[Dict, Dict]:
        """并行解析所有文件"""
        all_data, all_times = {}, {}
        
        for file, node in zip(result_files, node_names):
            try:
                all_data[node], all_times[node] = self._extract_data_from_file(file)
            except Exception as e:
                self.logger.warning(f"解析文件 {file} 失败: {e} -> 跳过")
                continue
                
        return all_data, all_times

    def _extract_data_from_file(self, filepath: Path) -> Tuple[Dict, Dict]:
        """从文件中提取数据"""
        data, all_times = OrderedDict(), OrderedDict()
        
        try:
            with filepath.open('r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    
                    if line.startswith('节点'):
                        self._parse_node_line(line, all_times)
                    else:
                        self._parse_time_range_line(line, data)
            
            return data, all_times
            
        except Exception as e:
            self.logger.error(f"读取文件 {filepath} 时出错: {e}")
            raise

    def _parse_node_line(self, line: str, all_times: Dict) -> None:
        """解析节点行"""
        try:
            parts = line.split(" - ")
            if len(parts) < 2:
                return
            
            start_part = parts[1].split('的所有时间戳：')
            if len(start_part) < 2:
                return
            
            start_key = start_part[0].strip()
            times = line.split("时间戳：")[-1].strip()
            all_times[start_key] = times
        except Exception:
            return
    
    def _parse_time_range_line(self, line: str, data: Dict) -> None:
        """解析时间范围行"""
        try:
            parts = line.split(" - ")
            if len(parts) < 2:
                return
            
            start_key = parts[0].strip()
            time_parts = parts[1].split("最早时间:")
            
            if len(time_parts) < 2:
                return
            
            shortest_part = time_parts[1].split(",")[0].strip()
            longest_part = parts[1].split("最晚时间:")[1].split("，")[0].strip()
            
            shortest = self._parse_time(shortest_part)
            longest = self._parse_time(longest_part)
            
            data[start_key] = (shortest, longest)
        except Exception:
            return

    def _parse_time(self, time_str: str) -> datetime:
        """解析时间字符串"""
        time_str = time_str.strip()
        
        # 尝试带毫秒的格式
        try:
            return datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            # 尝试不带毫秒的格式
            return datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")

    def _generate_output_files(self, logs_dir: Path, all_data: Dict, all_times: Dict) -> Dict[str, Path]:
        """生成所有输出文件"""
        result_dir = Path(logs_dir) / "res"
        if os.path.exists(result_dir):
            shutil.rmtree(result_dir)
        os.makedirs(result_dir, exist_ok=True)
        result_files = {
            'result_txt': result_dir / "result.txt",
            'result_excel': result_dir / "result.xlsx"
        }
        
        # # 清空simple_result_content
        # self.simple_result_content = []
        
        # 生成文本报告
        self._write_analysis_result(result_files['result_txt'], all_data, all_times)
        
        # 生成Excel报告（包含四个sheet）
        self._write_analysis_result_excel(result_files['result_excel'], all_data, all_times)
        
        return result_files

    def _write_analysis_result(self, result_file: Path, all_data: Dict, all_times: Dict) -> None:
        """写入分析结果"""
        try:
            with result_file.open('w', encoding='utf-8') as outfile:
                for i in range(len(self.start_keys) - 1):
                    start1, start2 = self.start_keys[i], self.start_keys[i + 1]
                    self._write_time_comparison(outfile, start1, start2, all_data, all_times)
                    
        except IOError as e:
            self.logger.error(f"写入结果文件失败: {e}")
            raise
    
    def _write_time_comparison(self, outfile, start1: str, start2: str, all_data: Dict, all_times: Dict) -> None:
        """写入时间比较结果"""
        threshold = 10
        
        start1_times = self._extract_times(all_data, start1, 0)
        start2_times = self._extract_times(all_data, start2, 1)
        
        if not start1_times or not start2_times:
            return
        
        # 构建输出内容
        output_lines = []
        output_lines.append(f"\n{start1}->{start2}\n")
        
        # 写入各节点耗时信息
        output_lines.append(self._format_node_times(f"【{start1}】各节点全部时间戳", start1, start1_times, all_times))
        output_lines.append(self._format_node_times(f"【{start2}】各节点全部时间戳", start2, start2_times, all_times))
        
        # 分析节点内耗时和等待时间
        all_node_start2_times, has_exceed_threshold = self._analyze_node_internal_times(
            output_lines, start1, start2, all_times, start1_times, start2_times
        )
        
        # 分析跨节点耗时
        max_diff, best_nodes = self._find_max_time_difference(start1_times, start2_times)
        
        if max_diff is not None:
            node1, node2 = best_nodes
            output_lines.append(f'-----------------------------【{start1}】-【{start2}】各节点间的执行耗时分析结果-----------------------------\n')
            output_lines.append(f"{start1} - {start2} 过程中节点 {node1} - {node2} 最大耗时: {max_diff}s\n")
            output_lines.append(f"  {start1} 最早时间 @ {node1}: {start1_times[node1]}\n")
            output_lines.append(f"  {start2} 最晚时间 @ {node2}: {start2_times[node2]}\n")
            
            if max_diff > threshold:
                self.gt_threshold_times.append(f"{start1}-{start2} 耗时最大节点: {node1} -> {node2}, 耗时: {max_diff}\n")
            
            self._write_waiting_analysis(output_lines, node2, start2_times[node2], all_node_start2_times)
        
        # 写入到result.txt
        for line in output_lines:
            outfile.write(line)
        
        # # 如果有节点内耗时超过阈值，则添加到simple_result_content
        # if has_exceed_threshold:
        #     for line in output_lines:
        #         self.simple_result_content.append(line)

    def _analyze_node_internal_times(self, output_lines: List[str], start1: str, start2: str, all_times: Dict, 
                                   start1_times: Dict, start2_times: Dict) -> Tuple[Dict, bool]:
        """分析节点内耗时，返回所有节点的start2最晚时间和是否有节点内耗时超过阈值"""
        output_lines.append(f'-----------------------------【{start1}】-【{start2}】各节点内的执行耗时分析结果-----------------------------\n')
        
        today, all_node_start2_times = date.today(), {}
        common_nodes = set(start1_times.keys()) & set(start2_times.keys())
        has_exceed_threshold = False
        
        for node in common_nodes:
            if node in all_times and start1 in all_times[node] and start2 in all_times[node]:
                min_start1 = min(self._parse_time_list(all_times[node][start1]))
                max_start2 = max(self._parse_time_list(all_times[node][start2]))
                all_node_start2_times[node] = max_start2
                
                node_diff = datetime.combine(today, max_start2) - datetime.combine(today, min_start1)
                node_diff_seconds = node_diff.total_seconds()
                
                # 写入节点内耗时信息
                output_lines.append(f"  节点{node}内耗时: {node_diff}\n")
                
                # 检查是否超过阈值
                if node_diff_seconds > self.config.interval_time:
                    has_exceed_threshold = True
        
        return all_node_start2_times, has_exceed_threshold
    
    def _format_node_times(self, title: str, start_key: str, times_dict: Dict, all_times: Dict) -> str:
        """格式化节点时间信息为字符串"""
        result = f" -----------------------------{title}-----------------------------\n"
        
        node_infos = []
        for node in times_dict.keys():
            time_info = all_times.get(node, {}).get(start_key, "无记录")
            node_infos.append(f"节点{node}: {time_info}")
        
        result += "\n，".join(node_infos) + "\n"
        return result
    
    def _write_waiting_analysis(self, output_lines: List[str], slowest_node: str, slowest_time: datetime, all_node_start2_times: Dict) -> None:
        """写入等待时间分析"""
        output_lines.append(f'-----------------------------结束：各节点等待节点 {slowest_node} 的最晚 {slowest_time} 耗时-----------------------------\n')
        
        today, max_time = date.today(), slowest_time.time()
        for node, node_time in all_node_start2_times.items():
            if node != slowest_node:
                wait_time = datetime.combine(today, max_time) - datetime.combine(today, node_time)
                output_lines.append(f'  节点 {node} ，最晚于 {node_time} 执行完成，共等待 {wait_time}s \n')
    
    def _extract_times(self, all_data: Dict, start_key: str, time_index: int) -> Dict:
        """提取指定时间点的数据"""
        return {
            node: data[start_key][time_index]
            for node, data in all_data.items()
            if start_key in data and data[start_key][time_index] is not None
        }
    
    def _parse_time_list(self, time_str: str):
        """解析时间列表"""
        time_strings = re.findall(r"\d{2}:\d{2}:\d{2}\.\d{3}", time_str)
        return [datetime.strptime(t, "%H:%M:%S.%f").time() for t in time_strings]
    
    def _find_max_time_difference(self, start1_times: Dict, start2_times: Dict) -> Tuple[Optional[float], Optional[Tuple]]:
        """查找最大时间差异"""
        max_diff, best_nodes = None, None
        
        for node1, time1 in start1_times.items():
            for node2, time2 in start2_times.items():
                current_diff = (time2 - time1).total_seconds()
                if max_diff is None or current_diff > max_diff:
                    max_diff, best_nodes = current_diff, (node1, node2)
        
        return max_diff, best_nodes
    
    def _write_analysis_result_excel(self, result_file: Path, all_data: Dict, all_times: Dict) -> None:
        """写入分析结果到 Excel，包含五个sheet"""
        try:
            wb = Workbook()
            
            # 第一个sheet：详细结果
            ws1 = wb.active
            ws1.title = "详细结果"

            row_offset = 1

            for i in range(len(self.start_keys) - 1):
                start1, start2 = self.start_keys[i], self.start_keys[i + 1]
                used_rows = self._write_time_comparison_excel(ws1, row_offset, start1, start2, all_data, all_times, simple=False)
                
                if used_rows > 0:
                    row_offset += used_rows + 2

            # 应用格式化和样式
            self._apply_excel_formatting(ws1)
            
            # 第二个sheet：阈值过滤结果
            ws2 = wb.create_sheet("阈值过滤结果")
            self._write_threshold_filtered_sheet(ws2, all_data, all_times)
            
            
            # 第三个sheet：节点慢阶段分析
            ws3 = wb.create_sheet("节点慢阶段分析")
            self._write_slow_stage_analysis(ws3, all_data, all_times)

            # 第四个sheet：简化结果
            ws4 = wb.create_sheet("间隔过滤结果")
            
            row_offset = 1
            for i in range(len(self.start_keys) - 1):
                start1, start2 = self.start_keys[i], self.start_keys[i + 1]
                used_rows = self._write_time_comparison_excel(ws4, row_offset, start1, start2, all_data, all_times, simple=True)
                
                if used_rows > 0:
                    row_offset += used_rows + 2

            # 应用格式化和样式
            self._apply_excel_formatting(ws4)

            # 第五个sheet：节点性能分析
            ws5 = wb.create_sheet("节点性能分析")
            
            # 创建节点性能分析表
            performance_df = self._create_performance_analysis_sheet(all_data, all_times)
            
            # 将性能分析数据写入第四个sheet
            if not performance_df.empty:
                self._write_performance_sheet(ws5, performance_df)
            else:
                ws5.cell(row=1, column=1, value="无性能分析数据")
            
            wb.save(result_file)
            self.logger.info(f"Excel 写入成功: {result_file}")

        except Exception as e:
            self.logger.error(f"写入结果文件失败: {e}")
            raise


    def _write_threshold_filtered_sheet(self, ws, all_data: Dict, all_times: Dict) -> None:
        """写入阈值过滤结果sheet - 增强版变化检测"""
        row_offset = 1
        last_step_info = None  # 存储上一个步骤的信息
        relative_delay_data = {}  # 存储相对延迟数据用于生成图表
        
        for i in range(len(self.start_keys) - 1):
            start1, start2 = self.start_keys[i], self.start_keys[i + 1]
            
            # 提取当前步骤的数据
            current_step_info = self._extract_step_info_for_comparison(all_data, all_times, start1, start2)
            
            # 检查是否应该跳过当前步骤
            if self._should_skip_step_by_comparison(current_step_info, last_step_info):
                continue  # 跳过这个步骤
            
            # 写入数据，并获取过滤后的节点和时间信息
            used_rows, filtered_nodes, filtered_times = self._write_threshold_filtered_data_with_info(
                ws, row_offset, start1, start2, all_data, all_times)
            
            if used_rows > 0:
                # 使用过滤后的数据计算相对延迟
                self._collect_relative_delay_from_filtered_data(
                    relative_delay_data, start1, start2, filtered_nodes, filtered_times)
                
                row_offset += used_rows + 2
                # 更新上一个步骤的信息
                last_step_info = current_step_info

        # 生成相对延迟折线图
        if relative_delay_data:
            self._create_relative_delay_chart(ws, row_offset, relative_delay_data)
        
        self._apply_excel_formatting(ws)


    def _write_threshold_filtered_data_with_info(self, ws, start_row: int, start1: str, start2: str, 
                                            all_data: Dict, all_times: Dict) -> tuple:
        """写入阈值过滤数据，并返回过滤后的节点和时间信息"""
        start1_times = self._extract_times(all_data, start1, 0)
        start2_times = self._extract_times(all_data, start2, 1)
        
        if not start1_times or not start2_times:
            return 0, [], {}
        
        # 构建节点时间数据
        all_node_times = self._build_node_times_data(all_times, start1, start2, start1_times, start2_times)
        
        # 过滤节点：只保留有数据的节点
        nodes = [node for node, times in all_node_times.items() 
                if start1 in times and start2 in times and times[start1] and times[start2]]
        
        if not nodes:
            return 0, [], {}
        
        # 写入表头和数据，并获取过滤后的节点和时间
        actual_rows, filtered_nodes, filtered_times = self._write_threshold_data_optimized_with_info(
            ws, start_row, start1, start2, nodes, all_node_times)
        
        return actual_rows, filtered_nodes, filtered_times




    def _write_threshold_data_optimized_with_info(self, ws, start_row: int, start1: str, start2: str, 
                                            nodes: List[str], all_node_times: Dict) -> tuple:
        """优化后的阈值过滤数据写入，返回过滤后的节点和时间信息"""
        # 创建灰色填充
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # 收集所有四行的数据
        all_rows_data = []
        row_labels = [
            f"{start1} 最早",
            f"{start1} 最晚", 
            f"{start2} 最早",
            f"{start2} 最晚"
        ]
        
        # 第一步：收集每行的原始数据
        for row_idx in range(4):
            row_data = []
            for col, node in enumerate(nodes, start=3):
                times = all_node_times[node]
                time_list = times.get(start1 if row_idx <= 1 else start2, [])
                
                if row_idx == 0:  # start1最早
                    time_val = min(time_list) if time_list else None
                elif row_idx == 1:  # start1最晚
                    time_val = max(time_list) if time_list else None
                elif row_idx == 2:  # start2最早
                    time_val = min(time_list) if time_list else None
                else:  # start2最晚 (row_idx == 3)
                    time_val = max(time_list) if time_list else None
                    
                row_data.append((col, time_val, node))
            all_rows_data.append(row_data)
        
        # 第二步：过滤数据 - 每行只保留最小值和大于阈值的数据
        filtered_rows_data = []
        valid_row_indices = []  # 记录哪些行有有效数据
        
        for row_idx, row_data in enumerate(all_rows_data):
            # 过滤掉None值
            valid_times = [(col, time, node) for col, time, node in row_data if time is not None]
            if not valid_times:
                filtered_rows_data.append([])
                continue
                
            # 找到最小值
            min_time = min(valid_times, key=lambda x: x[1])[1]
            
            # 计算阈值 - 将时间转换为秒数进行计算
            min_time_seconds = self._time_to_total_seconds(min_time)
            threshold_time_seconds = min_time_seconds + self.config.filter_threshold_seconds
            
            # 确定要保留的数据：最小值 + 大于阈值的数据
            keep_data = []
            for col, time, node in valid_times:
                time_seconds = self._time_to_total_seconds(time)
                if time == min_time or time_seconds > threshold_time_seconds:
                    keep_data.append((col, time, node))
            
            # 统计重复值
            time_counts = {}
            for _, time, _ in keep_data:
                time_str = self._format_time_no_ms(time)  # 使用不带毫秒的格式化
                time_counts[time_str] = time_counts.get(time_str, 0) + 1
            
            # 如果这一行只保留最小值（没有大于阈值的数据），则删除这一行
            has_over_threshold = any(self._time_to_total_seconds(time) > threshold_time_seconds for _, time, _ in keep_data)
            if not has_over_threshold:
                filtered_rows_data.append([])
            else:
                filtered_rows_data.append(keep_data)
                valid_row_indices.append(row_idx)
        
        # 第三步：检查是否需要删除整个阶段对
        if all(len(row_data) == 0 for row_data in filtered_rows_data):
            return 0, [], {}  # 整个阶段对都不显示
        
        # 第四步：确定要保留的节点（至少在一行中有数据的节点）
        valid_nodes = set()
        for row_data in filtered_rows_data:
            for _, _, node in row_data:
                valid_nodes.add(node)
        
        # 如果没有有效节点，则不显示
        if not valid_nodes:
            return 0, [], {}
        
        # 重新映射列索引
        node_to_col = {}
        for col, node in enumerate(sorted(valid_nodes), start=3):
            node_to_col[node] = col
        
        # 第五步：写入表头
        # 合并单元格显示阶段对
        start_merge_row = start_row + 1
        end_merge_row = start_row + len(valid_row_indices)
        
        if start_merge_row <= end_merge_row:
            ws.merge_cells(start_row=start_merge_row, start_column=1, 
                        end_row=end_merge_row, end_column=1)
            
            cell = ws.cell(row=start_merge_row, column=1, value=f"{start1}->{start2}")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 写入节点名称表头
        for node, col in node_to_col.items():
            cell = ws.cell(row=start_row, column=col, value=node)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 第六步：写入过滤后的数据 - 只写入有数据的行
        current_data_row = 0
        filtered_node_times = {}  # 存储过滤后的节点时间数据
        
        for row_idx in range(4):
            if row_idx not in valid_row_indices:
                continue  # 跳过空行
                
            row_data = filtered_rows_data[row_idx]
            if not row_data:  # 这一行被删除了
                continue
                
            # 写入行标签
            ws.cell(row=start_row + current_data_row + 1, column=2, value=row_labels[row_idx])
            
            # 统计重复值
            time_counts = {}
            for _, time, _ in row_data:
                time_str = self._format_time_no_ms(time)  # 使用不带毫秒的格式化
                time_counts[time_str] = time_counts.get(time_str, 0) + 1
            
            # 找到最小值
            min_time = min([time for _, time, _ in row_data])
            
            # 写入数据并收集过滤后的时间
            for col, time, node in row_data:
                if node not in node_to_col:
                    continue
                    
                target_col = node_to_col[node]
                cell = ws.cell(row=start_row + current_data_row + 1, column=target_col, 
                            value=self._format_time_no_ms(time))  # 使用不带毫秒的格式化
                
                # 存储过滤后的时间数据
                if node not in filtered_node_times:
                    filtered_node_times[node] = {}
                
                if row_idx == 0:  # start1最早
                    filtered_node_times[node]['start1_early'] = time
                elif row_idx == 1:  # start1最晚
                    filtered_node_times[node]['start1_late'] = time
                elif row_idx == 2:  # start2最早
                    filtered_node_times[node]['start2_early'] = time
                else:  # start2最晚
                    filtered_node_times[node]['start2_late'] = time
                
                # 标记重复值（不包括最小值）
                time_str = self._format_time_no_ms(time)
                if time_counts.get(time_str, 0) > 1 and time != min_time:
                    cell.fill = gray_fill
                elif time == min_time:
                    cell.fill = self.green_fill
            
            current_data_row += 1
        
        # 第七步：合并相邻相同值的单元格
        self._merge_adjacent_cells_optimized(ws, start_row, node_to_col, filtered_rows_data, valid_row_indices)
        
        # 返回实际使用的行数：表头1行 + 数据行数，以及过滤后的节点和时间
        return 1 + len(valid_row_indices), list(valid_nodes), filtered_node_times


    def _collect_relative_delay_from_filtered_data(self, relative_delay_data: Dict, start1: str, start2: str, 
                                                filtered_nodes: List[str], filtered_times: Dict) -> None:
        """
        使用过滤后的数据计算相对延迟
        
        Args:
            relative_delay_data: 存储相对延迟数据的字典
            start1: 起始阶段
            start2: 结束阶段
            filtered_nodes: 过滤后的节点列表
            filtered_times: 过滤后的时间数据
        """
        step_name = f"{start1}->{start2}"
        
        # 找到该步骤中最快的节点（start2最晚时间最小的节点）
        fastest_node = None
        fastest_time = None
        
        for node in filtered_nodes:
            if node in filtered_times and 'start2_late' in filtered_times[node]:
                time_seconds = self._time_to_total_seconds(filtered_times[node]['start2_late'])
                if fastest_time is None or time_seconds < fastest_time:
                    fastest_time = time_seconds
                    fastest_node = node
        
        if fastest_node is None:
            return
        
        # 计算每个节点的相对延迟（秒）
        step_delays = {}
        for node in filtered_nodes:
            if node in filtered_times and 'start2_late' in filtered_times[node]:
                node_time_seconds = self._time_to_total_seconds(filtered_times[node]['start2_late'])
                delay_seconds = node_time_seconds - fastest_time
                step_delays[node] = max(0, delay_seconds)  # 确保延迟不为负数
        
        relative_delay_data[step_name] = step_delays



    def _create_relative_delay_chart(self, ws, start_row: int, relative_delay_data: Dict) -> None:
        """
        创建相对延迟折线图 or 堆积折线图（自动判断）
        """
        try:
            # 准备图表数据
            chart_data = self._prepare_chart_data(relative_delay_data)
            if not chart_data:
                return
            
            # 写入图表数据到Excel
            data_start_row = start_row + 2
            data_rows = self._write_chart_data(ws, data_start_row, chart_data)
            
            stages_count = len(relative_delay_data)
            nodes_count = len(chart_data['nodes'])

            # ===============================
            #   根据阶段数量/节点数量判断图表类型
            # ===============================
            if stages_count > nodes_count:
                chart = LineChart()  # 普通折线图
                chart_type = "normal"
            else:
                chart = LineChart()  # 堆积折线图
                chart.grouping = "stacked"
                chart_type = "stacked"

            # 图表标题与坐标轴
            chart.title = "节点相对延迟趋势图"
            chart.style = 10
            chart.y_axis.title = "相对延迟（秒）"
            chart.x_axis.title = "执行阶段"
            chart.height = 15
            chart.width = 25

            # 数据范围
            data_ref = Reference(
                ws,
                min_col=2, min_row=data_start_row,
                max_col=1 + nodes_count, max_row=data_start_row + stages_count
            )

            categories_ref = Reference(
                ws,
                min_col=1, min_row=data_start_row + 1,
                max_row=data_start_row + stages_count
            )

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories_ref)

            # 根据类型优化样式
            self._optimize_chart_style(chart, nodes_count, chart_type)

            # 图表位置
            chart_cell = f"A{data_start_row + stages_count + 2}"
            ws.add_chart(chart, chart_cell)

        except ImportError:
            print("警告: 无法导入图表库，跳过图表生成")
        except Exception as e:
            print(f"创建图表时出错: {e}")


    def _optimize_chart_style(self, chart, nodes_count: int, chart_type: str) -> None:
        """
        优化图表样式，根据 chart_type（normal / stacked）调整风格
        """

        # 配色：普通折线图使用鲜明颜色；心电图（堆积线）使用柔和色
        normal_colors = [
            'FF5733', '33FF57', '3357FF', 'F333FF', '33FFF3',
            'FF33A1', 'A133FF', '33A1FF', 'FF8C33', '8CFF33'
        ]

        stacked_colors = [
            '88AAFF', '88FFAA', 'FFAA88', 'CC88FF', '88FFEE',
            'FF88CC', 'AA88FF', '88CCFF', 'FFCC88', 'CCFF88'
        ]

        # ---------------------------
        # 普通折线图风格（无标记）
        # ---------------------------
        if chart_type == "normal":
            for i, series in enumerate(chart.series):
                if i < len(normal_colors):
                    try:
                        series.graphicalProperties.line.solidFill = normal_colors[i]
                    except AttributeError:
                        pass

                series.graphicalProperties.line.width = 20000  # 粗一些
                series.marker = None  # 无标记

        # ---------------------------
        # 堆积折线图（心电图风格）
        # ---------------------------
        elif chart_type == "stacked":
            for i, series in enumerate(chart.series):
                # 柔和透明色
                if i < len(stacked_colors):
                    try:
                        series.graphicalProperties.line.solidFill = stacked_colors[i]
                    except AttributeError:
                        pass

                # 心电图效果：细线（0.5pt）
                series.graphicalProperties.line.width = 5000  

                # 去掉标记
                series.marker = None

                # 半透明效果（如果 openpyxl 支持）
                try:
                    series.graphicalProperties.line.alpha = 50000  # 50% 透明度
                except:
                    pass

        # ---------------------------
        # 坐标轴配置
        # ---------------------------
        try:
            chart.y_axis.title.text = "相对延迟（秒）"
            chart.y_axis.tickLblPos = "nextTo"
            chart.x_axis.tickLblPos = "nextTo"
            chart.x_axis.tickLblRot = 45
        except:
            pass

        chart.y_axis.scaling.min = 0
        chart.legend.position = 'tr'
        chart.layout = Layout(
            manualLayout=ManualLayout(x=0, y=0.04, w=0.9, h=0.78)
        )

        # 图表边框与背景
        try:
            chart.graphicalProperties.line.solidFill = "404040"
            chart.graphicalProperties.line.width = 10000
            chart.plot_area.graphicalProperties.solidFill = "F8F8F8"
        except:
            pass






    def _prepare_chart_data(self, relative_delay_data: Dict) -> Dict:
        """
        准备图表数据
        
        Returns:
            Dict: 包含节点列表和各阶段延迟数据的字典
        """
        # 收集所有节点
        all_nodes = set()
        for step_delays in relative_delay_data.values():
            all_nodes.update(step_delays.keys())
        
        if not all_nodes:
            return {}
        
        # 按节点名称排序
        sorted_nodes = sorted(all_nodes)
        
        # 构建数据矩阵
        stages = list(relative_delay_data.keys())
        data_matrix = {node: [] for node in sorted_nodes}
        
        for stage in stages:
            step_delays = relative_delay_data[stage]
            for node in sorted_nodes:
                delay = step_delays.get(node, 0)  # 如果没有数据，延迟为0
                data_matrix[node].append(delay)
        
        return {
            'stages': stages,
            'nodes': sorted_nodes,
            'data': data_matrix
        }



    def _write_chart_data(self, ws, start_row: int, chart_data: Dict) -> int:
        """
        将图表数据写入Excel
        
        Returns:
            int: 写入的数据行数
        """
        stages = chart_data['stages']
        nodes = chart_data['nodes']
        data_matrix = chart_data['data']
        
        # 写入表头
        header_row = start_row
        ws.cell(row=header_row, column=1, value="阶段")
        
        for i, node in enumerate(nodes, start=2):
            ws.cell(row=header_row, column=i, value=node)
        
        # 写入数据
        for i, stage in enumerate(stages):
            data_row = start_row + i + 1
            ws.cell(row=data_row, column=1, value=stage)
            
            for j, node in enumerate(nodes, start=2):
                delay = data_matrix[node][i]
                ws.cell(row=data_row, column=j, value=delay)
        
        # 添加详细说明文字
        note_row = start_row + len(stages) + 1
        ws.cell(row=note_row, column=1, 
            value="说明：相对延迟 = 节点完成当前阶段的最晚时间 - 最快节点完成当前阶段的最晚时间")
        

        return len(stages) + 1  # 返回数据行数（包括表头）





    def _extract_step_info_for_comparison(self, all_data: Dict, all_times: Dict, start1: str, start2: str) -> Dict:
        """
        提取步骤信息用于比较
        
        Returns:
            Dict: 包含节点集合和每个节点的关键时间信息
        """
        start1_times = self._extract_times(all_data, start1, 0)
        start2_times = self._extract_times(all_data, start2, 1)
        
        if not start1_times or not start2_times:
            return {}
        
        # 构建节点时间数据
        all_node_times = self._build_node_times_data(all_times, start1, start2, start1_times, start2_times)
        
        # 过滤节点：只保留有数据的节点
        nodes = [node for node, times in all_node_times.items() 
                if start1 in times and start2 in times and times[start1] and times[start2]]
        
        if not nodes:
            return {}
        
        # 提取每个节点的关键时间信息
        node_time_info = {}
        for node in nodes:
            times = all_node_times[node]
            start1_early = min(times[start1]) if times[start1] else None
            start2_early = min(times[start2]) if times[start2] else None
            
            node_time_info[node] = {
                'start1_early': start1_early,
                'start2_early': start2_early,
                'interval': self._time_to_total_seconds(start2_early) - self._time_to_total_seconds(start1_early) 
                            if start1_early and start2_early else None
            }
        
        return {
            'start1': start1,
            'start2': start2,
            'nodes': set(nodes),  # 使用集合便于比较
            'node_time_info': node_time_info
        }

    def _should_skip_step_by_comparison(self, current_step_info: Dict, last_step_info: Dict) -> bool:
        """
        通过比较判断是否应该跳过当前步骤
        
        条件：
        1. 节点集合相同
        2. 每个节点的start1最早时间与上一个步骤的start2最早时间差异在1秒内
        3. 每个节点的当前步骤间隔（start2-start1）小于1秒
        """
        if not last_step_info or not current_step_info:
            return False
        
        # 条件1：节点集合相同
        if current_step_info['nodes'] != last_step_info['nodes']:
            return False
        
        current_nodes = current_step_info['nodes']
        current_time_info = current_step_info['node_time_info']
        last_time_info = last_step_info['node_time_info']
        
        # 检查每个节点是否满足条件
        for node in current_nodes:
            if node not in current_time_info or node not in last_time_info:
                return False
            
            current_info = current_time_info[node]
            last_info = last_time_info[node]
            
            # 条件2：当前步骤的start1最早与上一个步骤的start2最早差异在1秒内
            if (current_info['start1_early'] and last_info['start2_early']):
                time_diff = abs(self._time_to_total_seconds(current_info['start1_early']) - 
                            self._time_to_total_seconds(last_info['start2_early']))
                if time_diff > 1:  # 大于1秒
                    return False
            
            # 条件3：当前步骤的间隔小于1秒
            if current_info['interval'] is not None and current_info['interval'] >= 1:
                return False
        
        # 所有节点都满足条件，跳过当前步骤
        return True

    def _write_threshold_filtered_data(self, ws, start_row: int, start1: str, start2: str, 
                                    all_data: Dict, all_times: Dict) -> int:
        """写入阈值过滤数据"""
        start1_times = self._extract_times(all_data, start1, 0)
        start2_times = self._extract_times(all_data, start2, 1)
        
        if not start1_times or not start2_times:
            return 0
        
        # 构建节点时间数据
        all_node_times = self._build_node_times_data(all_times, start1, start2, start1_times, start2_times)
        
        # 过滤节点：只保留有数据的节点
        nodes = [node for node, times in all_node_times.items() 
                if start1 in times and start2 in times and times[start1] and times[start2]]
        
        if not nodes:
            return 0
        
        # 写入表头和数据
        actual_rows = self._write_threshold_data_optimized(ws, start_row, start1, start2, nodes, all_node_times)
        
        return actual_rows

    def _time_to_total_seconds(self, time_val):
        """将时间值转换为总秒数（秒级精度）"""
        if time_val is None:
            return None
        
        # 如果已经是数字，假设是秒
        if isinstance(time_val, (int, float)):
            # 如果数字很大，假设是毫秒，转换为秒
            if time_val > 1000000000:  # 假设大于这个值的是毫秒时间戳
                return time_val / 1000
            return time_val
        
        # 如果是datetime.time对象，转换为从0点开始的秒数
        if isinstance(time_val, time):
            total_seconds = time_val.hour * 3600 + time_val.minute * 60 + time_val.second
            return total_seconds
        
        # 如果是字符串，尝试解析
        if isinstance(time_val, str):
            try:
                # 尝试解析 HH:MM:SS 格式
                parts = time_val.split(':')
                if len(parts) == 3:
                    hours, minutes, seconds = int(parts[0]), int(parts[1]), int(parts[2])
                    return hours * 3600 + minutes * 60 + seconds
            except:
                pass
        
        # 其他类型无法处理
        return None





    def _write_threshold_data_optimized(self, ws, start_row: int, start1: str, start2: str, 
                                    nodes: List[str], all_node_times: Dict) -> int:
        """优化后的阈值过滤数据写入"""
        # 创建灰色填充
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # 收集所有四行的数据
        all_rows_data = []
        row_labels = [
            f"{start1} 最早",
            f"{start1} 最晚", 
            f"{start2} 最早",
            f"{start2} 最晚"
        ]
        
        # 第一步：收集每行的原始数据
        for row_idx in range(4):
            row_data = []
            for col, node in enumerate(nodes, start=3):
                times = all_node_times[node]
                time_list = times.get(start1 if row_idx <= 1 else start2, [])
                
                if row_idx == 0:  # start1最早
                    time_val = min(time_list) if time_list else None
                elif row_idx == 1:  # start1最晚
                    time_val = max(time_list) if time_list else None
                elif row_idx == 2:  # start2最早
                    time_val = min(time_list) if time_list else None
                else:  # start2最晚 (row_idx == 3)
                    time_val = max(time_list) if time_list else None
                    
                row_data.append((col, time_val, node))
            all_rows_data.append(row_data)
        
        # 第二步：过滤数据 - 每行只保留最小值和大于阈值的数据
        filtered_rows_data = []
        valid_row_indices = []  # 记录哪些行有有效数据
        
        for row_idx, row_data in enumerate(all_rows_data):
            # 过滤掉None值
            valid_times = [(col, time, node) for col, time, node in row_data if time is not None]
            if not valid_times:
                filtered_rows_data.append([])
                continue
                
            # 找到最小值
            min_time = min(valid_times, key=lambda x: x[1])[1]
            
            # 计算阈值 - 将时间转换为秒数进行计算
            min_time_seconds = self._time_to_total_seconds(min_time)
            threshold_time_seconds = min_time_seconds + self.config.filter_threshold_seconds
            
            # 确定要保留的数据：最小值 + 大于阈值的数据
            keep_data = []
            for col, time, node in valid_times:
                time_seconds = self._time_to_total_seconds(time)
                if time == min_time or time_seconds > threshold_time_seconds:
                    keep_data.append((col, time, node))
            
            # 统计重复值
            time_counts = {}
            for _, time, _ in keep_data:
                time_str = self._format_time_no_ms(time)  # 使用不带毫秒的格式化
                time_counts[time_str] = time_counts.get(time_str, 0) + 1
            
            # 如果这一行只保留最小值（没有大于阈值的数据），则删除这一行
            has_over_threshold = any(self._time_to_total_seconds(time) > threshold_time_seconds for _, time, _ in keep_data)
            if not has_over_threshold:
                filtered_rows_data.append([])
            else:
                filtered_rows_data.append(keep_data)
                valid_row_indices.append(row_idx)
        
        # 第三步：检查是否需要删除整个阶段对
        if all(len(row_data) == 0 for row_data in filtered_rows_data):
            return 0  # 整个阶段对都不显示
        
        # 第四步：确定要保留的节点（至少在一行中有数据的节点）
        valid_nodes = set()
        for row_data in filtered_rows_data:
            for _, _, node in row_data:
                valid_nodes.add(node)
        
        # 如果没有有效节点，则不显示
        if not valid_nodes:
            return 0
        
        # 重新映射列索引
        node_to_col = {}
        for col, node in enumerate(sorted(valid_nodes), start=3):
            node_to_col[node] = col
        
        # 第五步：写入表头
        # 合并单元格显示阶段对
        start_merge_row = start_row + 1
        end_merge_row = start_row + len(valid_row_indices)
        
        if start_merge_row <= end_merge_row:
            ws.merge_cells(start_row=start_merge_row, start_column=1, 
                        end_row=end_merge_row, end_column=1)
            
            cell = ws.cell(row=start_merge_row, column=1, value=f"{start1}->{start2}")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 写入节点名称表头
        for node, col in node_to_col.items():
            cell = ws.cell(row=start_row, column=col, value=node)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 第六步：写入过滤后的数据 - 只写入有数据的行
        current_data_row = 0
        for row_idx in range(4):
            if row_idx not in valid_row_indices:
                continue  # 跳过空行
                
            row_data = filtered_rows_data[row_idx]
            if not row_data:  # 这一行被删除了
                continue
                
            # 写入行标签
            ws.cell(row=start_row + current_data_row + 1, column=2, value=row_labels[row_idx])
            
            # 统计重复值
            time_counts = {}
            for _, time, _ in row_data:
                time_str = self._format_time_no_ms(time)  # 使用不带毫秒的格式化
                time_counts[time_str] = time_counts.get(time_str, 0) + 1
            
            # 找到最小值
            min_time = min([time for _, time, _ in row_data])
            
            # 写入数据
            for col, time, node in row_data:
                if node not in node_to_col:
                    continue
                    
                target_col = node_to_col[node]
                cell = ws.cell(row=start_row + current_data_row + 1, column=target_col, 
                            value=self._format_time_no_ms(time))  # 使用不带毫秒的格式化
                
                # 标记重复值（不包括最小值）
                time_str = self._format_time_no_ms(time)
                if time_counts.get(time_str, 0) > 1 and time != min_time:
                    cell.fill = gray_fill
                elif time == min_time:
                    cell.fill = self.green_fill
            
            current_data_row += 1
        
        # 第七步：合并相邻相同值的单元格
        self._merge_adjacent_cells_optimized(ws, start_row, node_to_col, filtered_rows_data, valid_row_indices)
        
        # 返回实际使用的行数：表头1行 + 数据行数
        return 1 + len(valid_row_indices)

    def _format_time_no_ms(self, time_val):
        """格式化时间，去掉毫秒部分"""
        if time_val is None:
            return ""
        
        # 如果是datetime.time对象，直接格式化为%H:%M:%S
        if isinstance(time_val, time):
            return time_val.strftime("%H:%M:%S")
        
        # 如果是数字（秒数），转换为时间格式
        if isinstance(time_val, (int, float)):
            # 转换为时分秒
            hours = int(time_val // 3600)
            minutes = int((time_val % 3600) // 60)
            seconds = int(time_val % 60)
            
            if hours > 0:
                return f"{hours}:{minutes:02d}:{seconds:02d}"
            elif minutes > 0:
                return f"{minutes}:{seconds:02d}"
            else:
                return f"{seconds}秒"
        
        # 其他类型，返回字符串表示
        return str(time_val)

    def _merge_adjacent_cells_optimized(self, ws, start_row: int, node_to_col: Dict, 
                                    filtered_rows_data: List, valid_row_indices: List) -> None:
        """优化后的合并相邻单元格逻辑"""
        # 只对有数据的行进行合并
        for node, col in node_to_col.items():
            # 收集该节点在所有行中的值
            cell_values = []
            for row_idx in valid_row_indices:
                row_data = filtered_rows_data[row_idx]
                value = None
                for c, time, n in row_data:
                    if n == node:
                        value = str(self._format_time_no_ms(time))
                        break
                cell_values.append(value)
            
            # 合并相邻相同值的单元格
            start_merge = None
            current_value = None
            
            for i, value in enumerate(cell_values):
                if value == current_value and value is not None:
                    # 继续合并
                    continue
                else:
                    # 结束当前合并
                    if start_merge is not None and i - 1 > start_merge:
                        ws.merge_cells(
                            start_row=start_row + start_merge + 1,
                            start_column=col,
                            end_row=start_row + i,
                            end_column=col
                        )
                    
                    # 开始新的合并
                    start_merge = i
                    current_value = value
            
            # 处理最后的合并
            if start_merge is not None and len(cell_values) - 1 > start_merge:
                ws.merge_cells(
                    start_row=start_row + start_merge + 1,
                    start_column=col,
                    end_row=start_row + len(cell_values),
                    end_column=col
                )

    def _add_time_delta(self, time_obj: time, delta_seconds: float) -> time:
        """给时间对象添加秒数增量"""
        if not time_obj:
            return None
        
        # 将时间转换为当天的时间戳
        base_datetime = datetime.combine(date.today(), time_obj)
        # 添加增量
        new_datetime = base_datetime + timedelta(seconds=delta_seconds)
        # 转换回时间对象
        return new_datetime.time()

    def _write_time_comparison_excel(self, ws, start_row: int, start1: str, start2: str, 
                                all_data: Dict, all_times: Dict, simple: bool) -> int:
        """将时间比较结果写入Excel"""
        start1_times = self._extract_times(all_data, start1, 0)
        start2_times = self._extract_times(all_data, start2, 1)
        
        if not start1_times or not start2_times:
            return 0
        
        # 构建节点时间数据
        all_node_times = self._build_node_times_data(all_times, start1, start2, start1_times, start2_times)
        
        # 简单模式过滤节点
        if simple:
            all_node_times = self._filter_nodes_by_threshold(all_node_times, start1, start2)
            if not all_node_times:
                return 0
        
        nodes = list(all_node_times.keys())
        
        # 写入表头和数据
        self._write_excel_header(ws, start_row, start1, start2, nodes)
        
        if simple:
            # 简单模式：应用时间合并逻辑
            used_rows = self._write_excel_data_simple(ws, start_row, start1, start2, nodes, all_node_times)
        else:
            # 详细模式：保持原有逻辑
            self._write_excel_data(ws, start_row, start1, start2, nodes, all_node_times)
            used_rows = 7
        
        return used_rows

    def _write_excel_header(self, ws, start_row: int, start1: str, start2: str, nodes: List[str]) -> None:
        """写入Excel表头"""
        cell = ws.cell(row=start_row, column=1, value=f"{start1} - {start2}")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col, node in enumerate(nodes, start=3):
            cell = ws.cell(row=start_row, column=col, value=node)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _write_excel_data(self, ws, start_row: int, start1: str, start2: str, 
                        nodes: List[str], all_node_times: Dict) -> None:
        """写入Excel数据"""
        labels = [
            f"{start1} 最早", f"{start1} 最晚",
            f"{start2} 最早", f"{start2} 最晚",
            f"节点内 {start2}-{start1}（s）",
            f"节点间 {start2}-{start1}（s）",
        ]
        
        # 在第一列前插入新列，并合并单元格显示 {start1}->{start2}
        if len(nodes) > 0:
            # 合并从 start_row+1 到 start_row+len(labels) 的单元格
            start_col = 1
            end_col = 1
            start_merge_row = start_row + 1
            end_merge_row = start_row + len(labels)
            
            if start_merge_row <= end_merge_row:
                ws.merge_cells(start_row=start_merge_row, start_column=start_col, 
                            end_row=end_merge_row, end_column=end_col)
                
                # 在合并的单元格中写入内容
                cell = ws.cell(row=start_merge_row, column=start_col, 
                            value=f"{start1}->{start2}")
                
                # 设置居中对齐
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 写入行标签（现在从第二列开始）
        for i, label in enumerate(labels, start=1):
            ws.cell(row=start_row + i, column=2, value=label)
        
        # 收集所有时间值用于标记最小/最大值
        time_values = {label: {} for label in ['t1_min', 't1_max', 't2_min', 't2_max']}
        
        # 用于计算节点间耗时
        all_t1_min = []
        all_t2_max = []
        earliest_node = None
        latest_node = None
        
        # 写入节点数据（列索引从3开始）
        for col, node in enumerate(nodes, start=3):
            times = all_node_times[node]
            t1_list, t2_list = times.get(start1, []), times.get(start2, [])
            
            t1_min = min(t1_list) if t1_list else None
            t1_max = max(t1_list) if t1_list else None
            t2_min = min(t2_list) if t2_list else None
            t2_max = max(t2_list) if t2_list else None
            
            # 存储时间值
            time_values['t1_min'][col] = t1_min
            time_values['t1_max'][col] = t1_max
            time_values['t2_min'][col] = t2_min
            time_values['t2_max'][col] = t2_max
            
            # 收集所有节点的时间用于计算节点间耗时
            if t1_min:
                all_t1_min.append((t1_min, node))
            if t2_max:
                all_t2_max.append((t2_max, node))
            
            # 写入时间数据
            ws.cell(row=start_row + 1, column=col, value=str(t1_min) if t1_min else "")
            ws.cell(row=start_row + 2, column=col, value=str(t1_max) if t1_max else "")
            ws.cell(row=start_row + 3, column=col, value=str(t2_min) if t2_min else "")
            ws.cell(row=start_row + 4, column=col, value=str(t2_max) if t2_max else "")
            
            # 计算并写入节点内耗时
            if t1_min and t2_max:
                today = date.today()
                delta = (datetime.combine(today, t2_max) - datetime.combine(today, t1_min)).total_seconds()
                cell = ws.cell(row=start_row + 5, column=col, value=delta)
                if delta > 10:
                    cell.fill = self.red_fill
        
        # 计算并写入节点间耗时
        if all_t1_min and all_t2_max:
            # 找到最早的start1时间和对应的节点
            global_min_t1, earliest_node = min(all_t1_min, key=lambda x: x[0])
            # 找到最晚的start2时间和对应的节点
            global_max_t2, latest_node = max(all_t2_max, key=lambda x: x[0])
            
            today = date.today()
            cross_node_delta = (datetime.combine(today, global_max_t2) - datetime.combine(today, global_min_t1)).total_seconds()
            
            # 创建详细的描述文本
            description = f"节点{earliest_node}(最早:{global_min_t1.strftime('%H:%M:%S.%f')[:-3]}) -> 节点{latest_node}(最晚:{global_max_t2.strftime('%H:%M:%S.%f')[:-3]}) 间隔:{cross_node_delta:.3f}s"
            
            # 合并单元格并写入节点间耗时
            start_col = 3
            end_col = len(nodes) + 2
            
            # 合并单元格
            if start_col <= end_col:
                ws.merge_cells(start_row=start_row + 6, start_column=start_col, 
                            end_row=start_row + 6, end_column=end_col)
                
                # 写入节点间耗时描述
                cell = ws.cell(row=start_row + 6, column=start_col, value=description)
                
                # 设置水平和垂直居中
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # 如果超过10秒，背景标红
                if cross_node_delta > 10:
                    cell.fill = self.red_fill
        
        # 标记最小/最大值
        self._mark_extreme_values(ws, start_row, time_values)

    def _mark_extreme_values(self, ws, start_row: int, time_values: Dict) -> None:
        """标记极值（最小/最大值）"""
        # 标记最小值（绿色）- 只标记有值的列
        for label, fill in [('t1_min', self.green_fill), ('t2_min', self.green_fill)]:
            if time_values[label]:
                # 过滤掉None值
                valid_values = [t for t in time_values[label].values() if t]
                if valid_values:
                    min_val = min(valid_values)
                    for col, val in time_values[label].items():
                        if val == min_val:
                            row_offset = 1 if label == 't1_min' else 3
                            ws.cell(row=start_row + row_offset, column=col).fill = fill
        
        # 标记最大值（黄色）- 只标记有值的列
        for label, fill in [('t1_max', self.yellow_fill), ('t2_max', self.yellow_fill)]:
            if time_values[label]:
                # 过滤掉None值
                valid_values = [t for t in time_values[label].values() if t]
                if valid_values:
                    max_val = max(valid_values)
                    for col, val in time_values[label].items():
                        if val == max_val:
                            row_offset = 2 if label == 't1_max' else 4
                            ws.cell(row=start_row + row_offset, column=col).fill = fill

    def _write_excel_data_simple(self, ws, start_row: int, start1: str, start2: str, 
                            nodes: List[str], all_node_times: Dict) -> int:
        """写入简化模式的Excel数据，合并秒级别一致的时间单元格"""
        labels = [
            f"{start1} 最早", f"{start1} 最晚",
            f"{start2} 最早", f"{start2} 最晚",
            f"节点内 {start2}-{start1}（s）",
            f"节点间 {start2}-{start1}（s）",
        ]
        
        # 在第一列前插入新列，并合并单元格显示 {start1}->{start2}
        if len(nodes) > 0:
            start_col = 1
            end_col = 1
            start_merge_row = start_row + 1
            end_merge_row = start_row + len(labels)
            
            if start_merge_row <= end_merge_row:
                ws.merge_cells(start_row=start_merge_row, start_column=start_col, 
                            end_row=end_merge_row, end_column=end_col)
                
                cell = ws.cell(row=start_merge_row, column=start_col, 
                            value=f"{start1}->{start2}")
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 写入行标签
        for i, label in enumerate(labels, start=1):
            ws.cell(row=start_row + i, column=2, value=label)
        
        # 收集时间数据用于合并
        time_data = {
            't1_min': {}, 't1_max': {}, 
            't2_min': {}, 't2_max': {},
            'intra_node': {}, 'cross_node': {}
        }
        
        all_t1_min = []
        all_t2_max = []
        earliest_node = None
        latest_node = None
        
        # 首先收集所有数据
        for col, node in enumerate(nodes, start=3):
            times = all_node_times[node]
            t1_list, t2_list = times.get(start1, []), times.get(start2, [])
            
            t1_min = min(t1_list) if t1_list else None
            t1_max = max(t1_list) if t1_list else None
            t2_min = min(t2_list) if t2_list else None
            t2_max = max(t2_list) if t2_list else None
            
            # 存储时间数据
            time_data['t1_min'][col] = t1_min
            time_data['t1_max'][col] = t1_max
            time_data['t2_min'][col] = t2_min
            time_data['t2_max'][col] = t2_max
            
            if t1_min:
                all_t1_min.append((t1_min, node))
            if t2_max:
                all_t2_max.append((t2_max, node))
            
            # 计算节点内耗时
            if t1_min and t2_max:
                today = date.today()
                delta = (datetime.combine(today, t2_max) - datetime.combine(today, t1_min)).total_seconds()
                time_data['intra_node'][col] = delta
        
        # 计算节点间耗时
        if all_t1_min and all_t2_max:
            global_min_t1, earliest_node = min(all_t1_min, key=lambda x: x[0])
            global_max_t2, latest_node = max(all_t2_max, key=lambda x: x[0])
            
            today = date.today()
            cross_node_delta = (datetime.combine(today, global_max_t2) - datetime.combine(today, global_min_t1)).total_seconds()
            time_data['cross_node']['value'] = cross_node_delta
            time_data['cross_node']['description'] = (
                f"节点{earliest_node}(最早:{global_min_t1.strftime('%H:%M:%S.%f')[:-3]}) -> "
                f"节点{latest_node}(最晚:{global_max_t2.strftime('%H:%M:%S.%f')[:-3]}) 间隔:{cross_node_delta:.3f}s"
            )
        
        # 写入数据并合并秒级别一致的单元格
        self._write_and_merge_time_data(ws, start_row, nodes, time_data)
        
        return 7

    def _write_and_merge_time_data(self, ws, start_row: int, nodes: List[str], time_data: Dict) -> None:
        """写入时间数据并合并秒级别一致的单元格"""
        # 定义时间类型和对应的行偏移量
        time_types = [
            ('t1_min', 1, "最早时间"),
            ('t1_max', 2, "最晚时间"), 
            ('t2_min', 3, "最早时间"),
            ('t2_max', 4, "最晚时间")
        ]
        
        # 处理每个时间类型
        for time_type, row_offset, time_desc in time_types:
            current_row = start_row + row_offset
            data_dict = time_data[time_type]
            
            if not data_dict:
                continue
                
            # 按秒级别分组数据
            second_groups = self._group_by_seconds(data_dict)
            
            # 写入并合并单元格
            self._write_merged_cells(ws, current_row, second_groups, time_desc)
        
        # 写入节点内耗时（第5行）
        intra_node_row = start_row + 5
        intra_data = time_data['intra_node']
        if intra_data:
            # 对耗时数据进行分组（按整数秒分组）
            duration_groups = self._group_by_duration(intra_data)
            self._write_merged_cells(ws, intra_node_row, duration_groups, "节点内耗时")
        
        # 写入节点间耗时（第6行）
        cross_node_row = start_row + 6
        cross_data = time_data['cross_node']
        if cross_data and 'description' in cross_data:
            # 合并所有单元格写入节点间耗时描述
            start_col = 3
            end_col = len(nodes) + 2
            
            if start_col <= end_col:
                ws.merge_cells(
                    start_row=cross_node_row, 
                    start_column=start_col, 
                    end_row=cross_node_row, 
                    end_column=end_col
                )
                
                cell = ws.cell(row=cross_node_row, column=start_col, 
                            value=cross_data['description'])
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                if cross_data['value'] > 10:
                    cell.fill = self.red_fill

    def _write_merged_cells(self, ws, row: int, groups: List[Tuple[str, List[int]]], 
                        data_type: str) -> None:
        """写入合并的单元格"""
        for group_key, columns in groups:
            if len(columns) == 1:
                # 单列，不合并
                col = columns[0]
                cell = ws.cell(row=row, column=col, value=group_key)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                # 多列，合并单元格
                start_col = min(columns)
                end_col = max(columns)
                
                ws.merge_cells(
                    start_row=row, 
                    start_column=start_col, 
                    end_row=row, 
                    end_column=end_col
                )
                
                cell = ws.cell(row=row, column=start_col, value=group_key)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 根据数据类型设置不同的背景色
                if data_type == "节点内耗时":
                    try:
                        duration = float(group_key.replace('s', ''))
                        if duration > 10:
                            cell.fill = self.red_fill
                        elif duration > 5:
                            cell.fill = self.yellow_fill
                    except ValueError:
                        pass

    def _group_by_duration(self, data_dict: Dict) -> List[Tuple[str, List[int]]]:
        """按耗时数据进行分组（按整数秒分组）"""
        duration_groups = {}
        
        for col, duration in data_dict.items():
            if duration is None:
                continue
                
            # 按整数秒分组
            int_seconds = int(duration)
            duration_key = f"{int_seconds}s"
            
            if duration_key not in duration_groups:
                duration_groups[duration_key] = []
            duration_groups[duration_key].append(col)
        
        # 按列顺序排序并确保连续性
        grouped_list = []
        current_group = None
        
        for col in sorted(data_dict.keys()):
            if data_dict[col] is None:
                continue
                
            int_seconds = int(data_dict[col])
            duration_key = f"{int_seconds}s"
            
            if current_group is None:
                current_group = (duration_key, [col])
            elif current_group[0] == duration_key:
                current_group[1].append(col)
            else:
                if len(current_group[1]) > 1:
                    grouped_list.append(current_group)
                else:
                    grouped_list.append((current_group[0], current_group[1]))
                current_group = (duration_key, [col])
        
        if current_group and len(current_group[1]) > 0:
            if len(current_group[1]) > 1:
                grouped_list.append(current_group)
            else:
                grouped_list.append((current_group[0], current_group[1]))
        
        return grouped_list

    def _group_by_seconds(self, data_dict: Dict) -> List[Tuple[str, List[int]]]:
        """按秒级别对时间数据进行分组"""
        # 创建秒级别的分组
        second_groups = {}
        
        for col, time_val in data_dict.items():
            if time_val is None:
                continue
                
            # 将时间转换为秒级别的字符串（忽略毫秒）
            seconds_key = time_val.strftime('%H:%M:%S')
            
            if seconds_key not in second_groups:
                second_groups[seconds_key] = []
            second_groups[seconds_key].append(col)
        
        # 按列顺序排序每个分组中的列索引
        for key in second_groups:
            second_groups[key].sort()
        
        # 转换为列表并确保列顺序连续
        grouped_list = []
        current_group = None
        
        for col in sorted(data_dict.keys()):
            if data_dict[col] is None:
                continue
                
            seconds_key = data_dict[col].strftime('%H:%M:%S')
            
            if current_group is None:
                current_group = (seconds_key, [col])
            elif current_group[0] == seconds_key:
                current_group[1].append(col)
            else:
                if len(current_group[1]) > 1:
                    grouped_list.append(current_group)
                else:
                    # 单列不合并
                    grouped_list.append((current_group[0], current_group[1]))
                current_group = (seconds_key, [col])
        
        # 添加最后一个分组
        if current_group and len(current_group[1]) > 0:
            if len(current_group[1]) > 1:
                grouped_list.append(current_group)
            else:
                grouped_list.append((current_group[0], current_group[1]))
        
        return grouped_list

    def _build_node_times_data(self, all_times: Dict, start1: str, start2: str, 
                             start1_times: Dict, start2_times: Dict) -> Dict:
        """构建节点时间数据"""
        all_node_times = {}
        
        for node in start1_times:
            if node in all_times and start1 in all_times[node]:
                all_node_times.setdefault(node, {})
                all_node_times[node][start1] = self._parse_time_list(all_times[node][start1])
        
        for node in start2_times:
            if node in all_times and start2 in all_times[node]:
                all_node_times.setdefault(node, {})
                all_node_times[node][start2] = self._parse_time_list(all_times[node][start2])
        
        return all_node_times
    
    def _filter_nodes_by_threshold(self, all_node_times: Dict, start1: str, start2: str) -> Dict:
        """根据阈值过滤节点"""
        today, filtered_nodes = date.today(), {}
        
        for node, times in all_node_times.items():
            t1_list = times.get(start1, [])
            t2_list = times.get(start2, [])
            
            if not t1_list or not t2_list:
                continue
            
            t1_min, t2_max = min(t1_list), max(t2_list)
            delta = (datetime.combine(today, t2_max) - datetime.combine(today, t1_min)).total_seconds()
            
            if delta > self.config.interval_time:
                filtered_nodes[node] = times
        
        return filtered_nodes

    def _apply_excel_formatting(self, ws) -> None:
        """应用Excel格式"""
        # 所有单元格居中
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 自动列宽
        for col in ws.columns:
            max_len = max([len(str(cell.value)) if cell.value else 0 for cell in col])
            ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    def _write_slow_stage_analysis(self, ws, all_data: Dict, all_times: Dict) -> None:
        """写入节点慢阶段分析"""
        # 表头 - 增加最后三列
        headers = [
            "节点", "慢阶段次数", "首个慢阶段", 
            "其他慢阶段的平均时间", "差值",
            "最慢阶段", "其他慢阶段的平均时间", "差值",
            "最后的慢阶段", "其他慢阶段的平均时间", "差值"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # 收集所有节点
        all_nodes = set(all_data.keys())
        
        # 存储每个节点的慢阶段信息，用于后续统计
        node_slow_stages_map = {}
        
        # 分析每个节点的慢阶段
        row = 2
        for node in all_nodes:
            slow_stages_info = self._analyze_slow_stages_for_node(node, all_data, all_times)
            
            if not slow_stages_info['slow_stages']:
                continue
            
            slow_stages = slow_stages_info['slow_stages']
            stage_avg_times = slow_stages_info['stage_avg_times']
            
            # 存储节点的慢阶段信息
            node_slow_stages_map[node] = slow_stages
            
            # 计算统计信息
            slow_stage_count = len(slow_stages)
            
            # 按阶段顺序排序
            sorted_stages = sorted(slow_stages.items(), 
                                 key=lambda x: self.start_keys.index(x[0]) if x[0] in self.start_keys else len(self.start_keys))
            
            # 首个慢阶段
            first_stage_name, first_stage_data = sorted_stages[0]
            first_stage_avg_time_seconds = stage_avg_times.get(first_stage_name, 0)
            first_stage_avg_time_str = self._seconds_to_time_str(first_stage_avg_time_seconds)
            first_stage_time_str = first_stage_data['node_time']
            first_stage_time = datetime.strptime(first_stage_time_str, "%H:%M:%S.%f").time()
            first_stage_seconds = (datetime.combine(date.today(), first_stage_time) - datetime.combine(date.today(), datetime.min.time())).total_seconds()
            first_stage_value = f"{first_stage_name}-{first_stage_time_str}"
            first_stage_diff = first_stage_seconds - first_stage_avg_time_seconds
            
            # 最慢阶段（差值最大的阶段）
            slowest_stage_name, slowest_stage_data = max(slow_stages.items(), 
                                                       key=lambda x: x[1]['time_diff'])
            slowest_stage_avg_time_seconds = stage_avg_times.get(slowest_stage_name, 0)
            slowest_stage_avg_time_str = self._seconds_to_time_str(slowest_stage_avg_time_seconds)
            slowest_stage_time_str = slowest_stage_data['node_time']
            slowest_stage_time = datetime.strptime(slowest_stage_time_str, "%H:%M:%S.%f").time()
            slowest_stage_seconds = (datetime.combine(date.today(), slowest_stage_time) - datetime.combine(date.today(), datetime.min.time())).total_seconds()
            slowest_stage_value = f"{slowest_stage_name}-{slowest_stage_time_str}"
            slowest_stage_diff = slowest_stage_seconds - slowest_stage_avg_time_seconds
            
            # 最后的慢阶段（按阶段顺序的最后一个）
            last_stage_name, last_stage_data = sorted_stages[-1]
            last_stage_avg_time_seconds = stage_avg_times.get(last_stage_name, 0)
            last_stage_avg_time_str = self._seconds_to_time_str(last_stage_avg_time_seconds)
            last_stage_time_str = last_stage_data['node_time']
            last_stage_time = datetime.strptime(last_stage_time_str, "%H:%M:%S.%f").time()
            last_stage_seconds = (datetime.combine(date.today(), last_stage_time) - datetime.combine(date.today(), datetime.min.time())).total_seconds()
            last_stage_value = f"{last_stage_name}-{last_stage_time_str}"
            last_stage_diff = last_stage_seconds - last_stage_avg_time_seconds
            
            # 写入数据
            ws.cell(row=row, column=1, value=node)
            ws.cell(row=row, column=2, value=slow_stage_count)
            ws.cell(row=row, column=3, value=first_stage_value)
            ws.cell(row=row, column=4, value=first_stage_avg_time_str)
            ws.cell(row=row, column=5, value=f"{first_stage_diff:.3f}s")
            ws.cell(row=row, column=6, value=slowest_stage_value)
            ws.cell(row=row, column=7, value=slowest_stage_avg_time_str)
            ws.cell(row=row, column=8, value=f"{slowest_stage_diff:.3f}s")
            # 新增最后三列
            ws.cell(row=row, column=9, value=last_stage_value)
            ws.cell(row=row, column=10, value=last_stage_avg_time_str)
            ws.cell(row=row, column=11, value=f"{last_stage_diff:.3f}s")
            
            # 如果慢阶段次数较多，标记为黄色
            if slow_stage_count > 3:
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = self.yellow_fill
            
            row += 1
        
        if row == 2:  # 没有数据
            ws.cell(row=2, column=1, value="无慢阶段数据")
        else:
            # 空6行
            row += 6
            
            # 添加新的表格：节点所有慢阶段统计
            self._write_slow_stage_statistics(ws, row, all_nodes, node_slow_stages_map)
        
        # 设置列宽和格式
        column_widths = [12, 12, 15, 18, 10, 15, 18, 10, 15, 18, 10]
        for col, width in enumerate(column_widths, 1):
            column_letter = self._get_column_letter(col)
            ws.column_dimensions[column_letter].width = width
        
        # 所有单元格居中
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

                
    def _write_slow_stage_statistics(self, ws, start_row: int, all_nodes: List[str], node_slow_stages_map: Dict) -> None:
        """写入节点所有慢阶段统计表格 - 只包含有慢阶段数据的节点，合并相邻且时间相同的慢阶段"""
        # 只选择有慢阶段数据的节点
        slow_nodes = [node for node in all_nodes if node in node_slow_stages_map and node_slow_stages_map[node]]
        
        # 如果没有慢节点，则不写入任何内容
        if not slow_nodes:
            return
        
        # 表头 - 使用合并后的阶段名
        headers = ["节点"] + self._get_merged_stage_headers(node_slow_stages_map, slow_nodes)
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        # 写入数据 - 只写入有慢阶段数据的节点
        row = start_row + 1
        for node in sorted(slow_nodes):
            ws.cell(row=row, column=1, value=node)
            
            # 获取该节点的合并慢阶段信息
            merged_stages = self._get_merged_slow_stages_for_node(node, node_slow_stages_map[node])
            
            # 对于每个合并后的阶段，检查该节点是否为慢阶段
            for col, stage_info in enumerate(headers[1:], 2):  # 跳过"节点"列
                stage_name = stage_info
                # 检查这个合并阶段是否包含该节点的慢阶段
                if stage_name in merged_stages:
                    # 是慢阶段，显示该阶段的最晚时间
                    node_time = merged_stages[stage_name]['node_time']
                    time_diff = merged_stages[stage_name]['time_diff']
                    cell = ws.cell(row=row, column=col, value=node_time)
                    
                    # 根据时间差大小设置不同颜色
                    if time_diff > 5.0:
                        cell.fill = self.red_fill
                    elif time_diff > 2.0:
                        cell.fill = self.yellow_fill
                    else:
                        cell.fill = self.green_fill
                else:
                    # 不是慢阶段，留空
                    ws.cell(row=row, column=col, value="")
            
            row += 1
        
        # 设置列宽
        for col in range(1, len(headers) + 1):
            column_letter = self._get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
        
        # 所有单元格居中
        for row_num in range(start_row, row):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center', vertical='center')

    def _get_merged_stage_headers(self, node_slow_stages_map: Dict, slow_nodes: List[str]) -> List[str]:
        """获取合并后的阶段表头"""
        # 收集所有节点中的所有慢阶段
        all_slow_stages = set()
        for node in slow_nodes:
            if node in node_slow_stages_map:
                all_slow_stages.update(node_slow_stages_map[node].keys())
        
        # 如果没有慢阶段，返回空列表
        if not all_slow_stages:
            return []
        
        # 将阶段按原始顺序排序
        sorted_stages = [stage for stage in self.start_keys if stage in all_slow_stages]
        
        # 合并相邻且时间相同的阶段
        merged_headers = []
        i = 0
        while i < len(sorted_stages):
            current_stage = sorted_stages[i]
            
            # 检查这个阶段在所有节点中的时间是否一致
            consistent_time = True
            stage_time = None
            
            for node in slow_nodes:
                if node in node_slow_stages_map and current_stage in node_slow_stages_map[node]:
                    if stage_time is None:
                        stage_time = node_slow_stages_map[node][current_stage]['node_time']
                    elif node_slow_stages_map[node][current_stage]['node_time'] != stage_time:
                        consistent_time = False
                        break
            
            # 如果时间不一致，直接添加当前阶段
            if not consistent_time:
                merged_headers.append(current_stage)
                i += 1
                continue
            
            # 查找连续的相同时间阶段
            j = i + 1
            while j < len(sorted_stages):
                next_stage = sorted_stages[j]
                
                # 检查下一个阶段是否连续且时间相同
                consistent_next = True
                for node in slow_nodes:
                    if (node in node_slow_stages_map and 
                        next_stage in node_slow_stages_map[node] and
                        node_slow_stages_map[node][next_stage]['node_time'] != stage_time):
                        consistent_next = False
                        break
                
                if consistent_next:
                    j += 1
                else:
                    break
            
            # 如果有连续阶段，合并它们
            if j > i + 1:
                merged_header = "->".join(sorted_stages[i:j])
                merged_headers.append(merged_header)
                i = j
            else:
                merged_headers.append(current_stage)
                i += 1
        
        return merged_headers

    def _get_merged_slow_stages_for_node(self, node: str, node_slow_stages: Dict) -> Dict:
        """获取节点合并后的慢阶段信息"""
        if not node_slow_stages:
            return {}
        
        # 将阶段按原始顺序排序
        sorted_stages = [stage for stage in self.start_keys if stage in node_slow_stages]
        
        # 合并相邻且时间相同的阶段
        merged_stages = {}
        i = 0
        while i < len(sorted_stages):
            current_stage = sorted_stages[i]
            current_time = node_slow_stages[current_stage]['node_time']
            
            # 查找连续的相同时间阶段
            j = i + 1
            while j < len(sorted_stages):
                next_stage = sorted_stages[j]
                if (next_stage in node_slow_stages and 
                    node_slow_stages[next_stage]['node_time'] == current_time):
                    j += 1
                else:
                    break
            
            # 如果有连续阶段，合并它们
            if j > i + 1:
                merged_stage_name = "->".join(sorted_stages[i:j])
                # 使用最后一个阶段的时间差作为合并阶段的时间差
                last_stage = sorted_stages[j-1]
                merged_stages[merged_stage_name] = {
                    'node_time': current_time,
                    'time_diff': node_slow_stages[last_stage]['time_diff']
                }
                i = j
            else:
                merged_stages[current_stage] = node_slow_stages[current_stage]
                i += 1
        
        return merged_stages





    def _get_column_letter(self, col_idx: int) -> str:
        """将列索引转换为Excel列名"""
        letters = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            letters = chr(65 + remainder) + letters
        return letters

    def _seconds_to_time_str(self, total_seconds: float) -> str:
        """将总秒数转换为时分秒格式"""
        if total_seconds <= 0:
            return "00:00:00.000"
        
        # 计算小时、分钟、秒和毫秒
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = total_seconds % 60
        milliseconds = int((seconds - int(seconds)) * 1000)
        
        # 格式化为 HH:MM:SS.sss
        return f"{hours:02d}:{minutes:02d}:{int(seconds):02d}.{milliseconds:03d}"

    def _analyze_slow_stages_for_node(self, node: str, all_data: Dict, all_times: Dict) -> Dict:
        """分析单个节点的慢阶段"""
        slow_stages = {}
        stage_avg_times = {}
        
        # 首先计算每个阶段所有节点的平均最晚时间（以秒为单位）
        for start_key in self.start_keys:
            # 收集所有节点在该阶段的最晚时间
            stage_times = []
            for other_node, node_data in all_data.items():
                if start_key in node_data:
                    # 获取最晚时间并转换为秒数
                    latest_time = node_data[start_key][1]
                    # 将datetime转换为当天的秒数
                    time_seconds = (latest_time - datetime.combine(latest_time.date(), datetime.min.time())).total_seconds()
                    stage_times.append(time_seconds)
            
            if stage_times:
                # 计算该阶段所有节点的平均最晚时间（秒）
                avg_time_seconds = sum(stage_times) / len(stage_times)
                stage_avg_times[start_key] = avg_time_seconds
        
        # 分析当前节点的慢阶段
        for start_key in self.start_keys:
            # 检查节点是否有该阶段的数据
            if node not in all_data or start_key not in all_data[node]:
                continue
            
            # 获取当前节点在该阶段的最晚时间
            node_time = all_data[node][start_key][1]
            
            # 将节点时间转换为秒数
            node_time_seconds = (node_time - datetime.combine(node_time.date(), datetime.min.time())).total_seconds()
            
            # 获取其他所有节点在该阶段的最晚时间（秒数）
            other_nodes_times = []
            for other_node, node_data in all_data.items():
                if other_node != node and start_key in node_data:
                    other_time = node_data[start_key][1]
                    other_time_seconds = (other_time - datetime.combine(other_time.date(), datetime.min.time())).total_seconds()
                    other_nodes_times.append(other_time_seconds)
            
            if not other_nodes_times:
                continue
            
            # 计算其他节点的最大最晚时间（秒数）
            max_other_time = max(other_nodes_times)
            
            # 检查当前节点是否比其他节点慢（超过阈值）
            time_diff = node_time_seconds - max_other_time
            
            if time_diff > self.config.slow_stage_threshold:
                # 格式化节点时间为字符串
                node_time_str = node_time.strftime('%H:%M:%S.%f')[:-3]
                
                slow_stages[start_key] = {
                    'node_time': node_time_str,
                    'max_other_time': max_other_time,
                    'time_diff': time_diff
                }
        
        return {
            'slow_stages': slow_stages,
            'stage_avg_times': stage_avg_times
        }

    def _create_performance_analysis_sheet(self, all_data: Dict, all_times: Dict) -> pd.DataFrame:
        """创建节点性能分析表"""
        # 提取所有节点
        nodes = list(all_data.keys())
        if not nodes:
            return pd.DataFrame()
        
        # 使用列表推导式一次性创建所有数据
        performance_data = []
        
        for node in nodes:
            node_perf = {'节点名称': node}
            
            # 为每个阶段对计算性能指标
            stage_durations = {}
            for i in range(len(self.start_keys) - 1):
                start1, start2 = self.start_keys[i], self.start_keys[i + 1]
                
                # 获取时间数据
                start1_times = self._extract_times(all_data, start1, 0)
                start2_times = self._extract_times(all_data, start2, 1)
                
                if node not in start1_times or node not in start2_times:
                    continue
                    
                # 计算节点内耗时
                t1_min = start1_times[node]
                t2_max = start2_times[node]
                intra_node_duration = (t2_max - t1_min).total_seconds()
                
                # 使用原始阶段对名称
                stage_pair_name = f"{start1}→{start2}"
                stage_durations[stage_pair_name] = intra_node_duration
            
            # 添加阶段耗时数据
            node_perf.update(stage_durations)
            performance_data.append(node_perf)
        
        # 一次性创建DataFrame
        df_perf = pd.DataFrame(performance_data)
        
        # 计算全局性能指标
        if not df_perf.empty:
            df_perf = self._calculate_performance_metrics(df_perf)
        
        return df_perf

    def _calculate_performance_metrics(self, df_perf: pd.DataFrame) -> pd.DataFrame:
        """计算性能指标"""
        # 获取所有阶段耗时列（排除节点名称列）
        duration_cols = [col for col in df_perf.columns if col != '节点名称' and '→' in col]
        
        if not duration_cols:
            return df_perf
        
        # 创建一个新的DataFrame来存储所有计算结果
        new_columns = {}
        
        # 数据质量检查和预处理
        valid_duration_cols = []
        
        for col in duration_cols:
            # 检查该列是否全为0
            if (df_perf[col] == 0).all():
                continue
                
            valid_duration_cols.append(col)
        
        # 如果没有有效的阶段列，返回原始数据
        if not valid_duration_cols:
            # 添加默认的基础列
            new_columns['综合性能评分'] = 50
            new_columns['性能等级'] = '中等'
            new_columns['瓶颈标记'] = False
            new_columns['所有间隔内代码执行总耗时（s）'] = 0
            new_columns['综合耗时排名'] = 1
            new_columns['结果分析'] = "所有阶段耗时均为0，无法进行性能分析"
            
            # 使用pd.concat一次性添加所有新列
            new_columns_df = pd.DataFrame(new_columns, index=df_perf.index)
            df_perf = pd.concat([df_perf, new_columns_df], axis=1)
            return df_perf
        
        # 创建包含节点名称的处理后数据
        processed_data = df_perf[['节点名称'] + valid_duration_cols].copy()
        
        # 对部分为0的节点，统一替换为1e-9
        for col in valid_duration_cols:
            # 将该列中的0替换为1e-9
            processed_data[col] = processed_data[col].replace(0, 1e-9)
        
        # 筛选出有意义的阶段（最大值大于1秒的阶段）
        meaningful_stages = []
        for col in valid_duration_cols:
            if processed_data[col].max() > 1.0:
                meaningful_stages.append(col)
        
        # 如果没有有意义的阶段，回退到使用所有有效阶段
        stages_for_scoring = meaningful_stages if meaningful_stages else valid_duration_cols
        
        # 计算百分位列（耗时越短，百分位越高）- 使用选定的阶段
        percentile_data = {}
        for col in stages_for_scoring:
            try:
                percentile_data[f'{col}_百分位'] = (1 - processed_data[col].rank(pct=True)) * 100
            except Exception:
                # 如果出错，设置为默认值
                percentile_data[f'{col}_百分位'] = 50
        
        # 将百分位数据添加到新列字典
        new_columns.update(percentile_data)
        
        # 计算综合性能评分（所有选定阶段百分位的平均值）
        if stages_for_scoring:
            percentile_cols = [f'{col}_百分位' for col in stages_for_scoring]
            percentile_df = pd.DataFrame(percentile_data, index=df_perf.index)
            new_columns['综合性能评分'] = percentile_df.mean(axis=1).round(1)
        else:
            # 如果没有阶段，设置默认评分
            new_columns['综合性能评分'] = 50
        
        # 计算性能等级
        conditions = [
            new_columns['综合性能评分'] >= 80,
            new_columns['综合性能评分'] >= 60,
            new_columns['综合性能评分'] >= 40
        ]
        choices = ['优秀', '良好', '中等', '待优化']
        new_columns['性能等级'] = np.select(conditions, choices[:3], default=choices[3])
        
        # 计算每个节点的总耗时（使用原始数据，但只计算有效列）
        total_durations = df_perf[valid_duration_cols].sum(axis=1)
        
        # 修改瓶颈标记逻辑：两个步骤，使用选定的阶段
        bottleneck_nodes = set()
        
        # 步骤1：标记在超过50%的选定阶段都是慢节点的节点
        bottleneck_scores = {}
        for col in stages_for_scoring:
            try:
                # 统一使用80%分位数作为阈值
                threshold = processed_data[col].quantile(0.8)
                # 使用包含节点名称的processed_data来筛选
                slow_nodes = processed_data[processed_data[col] >= threshold]['节点名称'].tolist()
                for node in slow_nodes:
                    bottleneck_scores[node] = bottleneck_scores.get(node, 0) + 1
            except Exception:
                pass
        
        # 在超过50%选定阶段都是慢节点的标记为瓶颈
        bottleneck_threshold = len(stages_for_scoring) * 0.5
        bottleneck_nodes_step1 = [node for node, score in bottleneck_scores.items() 
                        if score >= bottleneck_threshold]
        bottleneck_nodes.update(bottleneck_nodes_step1)
        
        # 步骤2：标记总延迟（各阶段延迟之和）在所有节点中排名前20%的节点
        # 使用选定阶段的总延迟
        try:
            # 计算选定阶段的总延迟
            scoring_total_durations = df_perf[stages_for_scoring].sum(axis=1)
            # 计算总延迟的80%分位数作为阈值（排名前20%）
            total_duration_threshold = scoring_total_durations.quantile(0.8)
            # 获取总延迟大于等于阈值的节点名称
            bottleneck_nodes_step2 = df_perf.loc[scoring_total_durations >= total_duration_threshold, '节点名称'].tolist()
            bottleneck_nodes.update(bottleneck_nodes_step2)
        except Exception:
            # 如果计算失败，跳过步骤2
            pass
        
        new_columns['瓶颈标记'] = df_perf['节点名称'].isin(bottleneck_nodes)
        
        # 添加所有阶段总耗时（s）列
        new_columns['所有间隔内代码执行总耗时（s）'] = total_durations.round(3)
        
        # 综合耗时排名（使用所有有效阶段的总耗时）
        new_columns['综合耗时排名'] = total_durations.rank(method='min', ascending=True).astype(int)
        
        # 添加结果分析列
        new_columns['结果分析'] = self._generate_performance_conclusions(
            new_columns['综合性能评分'],
            new_columns['性能等级'], 
            new_columns['瓶颈标记'],
            new_columns['综合耗时排名'],
            len(df_perf),
            len(stages_for_scoring),  # 修改：使用选定的阶段数量
            len(duration_cols)
        )
        
        # 使用pd.concat一次性添加所有新列，确保索引对齐
        new_columns_df = pd.DataFrame(new_columns, index=df_perf.index)
        df_perf = pd.concat([df_perf, new_columns_df], axis=1)
        
        # 重新排列列顺序：节点名称 + 综合指标 + 阶段耗时
        base_cols = ['节点名称', '综合性能评分', '性能等级', '瓶颈标记', 
                    '所有间隔内代码执行总耗时（s）', '综合耗时排名', '结果分析']
        # 只保留有效的阶段列
        stage_cols = [col for col in valid_duration_cols if col in df_perf.columns]
        stage_cols += [f'{col}_百分位' for col in stages_for_scoring if f'{col}_百分位' in df_perf.columns]
        
        # 确保所有列都存在
        final_cols = base_cols + stage_cols
        df_perf = df_perf.reindex(columns=[col for col in final_cols if col in df_perf.columns])
        
        return df_perf



    def _generate_performance_conclusions(self, performance_scores, performance_levels, 
                                    bottleneck_flags, 
                                    performance_ranks, total_nodes, 
                                    valid_stages_count, total_stages_count):
        """生成性能分析结论"""
        conclusions = []
        
        for i in range(len(performance_scores)):
            score = performance_scores[i]
            level = performance_levels[i]
            is_bottleneck = bottleneck_flags[i]
            rank = performance_ranks[i]
            
            conclusion_parts = []
            
            # 0. 阶段有效性信息
            if valid_stages_count < total_stages_count:
                conclusion_parts.append(f"基于{valid_stages_count}/{total_stages_count}个有效阶段")
            
            # 1. 基础性能评估
            if level == '优秀':
                conclusion_parts.append("性能优秀")
            elif level == '良好':
                conclusion_parts.append("性能良好")
            elif level == '中等':
                conclusion_parts.append("性能中等，有优化空间")
            else:
                conclusion_parts.append("性能较差，需重点优化")
            
            # 2. 瓶颈识别
            if is_bottleneck:
                conclusion_parts.append("系统瓶颈节点")
            
            # 3. 排名评估
            rank_percentile = (rank / total_nodes) * 100
            if rank_percentile <= 20:
                conclusion_parts.append("排名优秀")
            elif rank_percentile <= 40:
                conclusion_parts.append("排名良好")
            elif rank_percentile <= 60:
                conclusion_parts.append("排名中等")
            elif rank_percentile <= 80:
                conclusion_parts.append("排名较差")
            else:
                conclusion_parts.append("排名差")
            
            # 4. 特殊问题识别
            if level in ['中等', '待优化'] and is_bottleneck:
                conclusion_parts.append("性能与瓶颈均需改善")
            elif is_bottleneck and rank_percentile > 80:
                conclusion_parts.append("瓶颈且整体性能差")
            
            conclusions.append("；".join(conclusion_parts))
        
        return conclusions

    def _write_performance_sheet(self, ws, performance_df: pd.DataFrame) -> None:
        """将性能分析数据写入sheet"""
        from openpyxl.utils import get_column_letter
        
        # 基础列配置
        base_cols = ['节点名称', '综合性能评分', '性能等级', '瓶颈标记', 
                    '所有间隔内代码执行总耗时（s）', '综合耗时排名', '结果分析']
        
        # 阶段列（包含耗时和百分位）- 只保留有效的阶段列
        stage_cols = [col for col in performance_df.columns if col not in base_cols and '→' in col and '_百分位' not in col]
        percentile_cols = [col for col in performance_df.columns if col not in base_cols and '_百分位' in col]
        
        # ========== 写入数据 ==========
        # 第一行表头 - 基础列 + 阶段耗时
        headers_row1 = []
        
        # 基础列
        for col in base_cols:
            headers_row1.append(col)
        
        # 阶段耗时列
        for stage_col in stage_cols:
            headers_row1.append(f'{stage_col}耗时')
        
        # 写入第一行表头
        for col_idx, header in enumerate(headers_row1, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # 写入节点耗时数据（第2行开始）
        for row_idx, (_, row_data) in enumerate(performance_df.iterrows(), 2):
            # 写入基础列数据
            for col_idx, col_name in enumerate(base_cols, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
                
                # 为关键列设置不同的对齐方式
                if col_name == '结果分析':
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 为关键列添加特殊背景色
                if col_name == '所有阶段总耗时（s）':
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            # 写入阶段耗时数据
            for col_idx, stage_col in enumerate(stage_cols, len(base_cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[stage_col])
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        
        # 计算百分位数据开始的行
        start_percentile_row = len(performance_df) + 3
        
        # 写入百分位表头（在所有节点耗时数据后添加）
        # 基础列位置留空
        for col_idx in range(1, len(base_cols) + 1):
            cell = ws.cell(row=start_percentile_row, column=col_idx, value='')
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # 阶段百分位列标题 - 只写入实际存在的百分位列
        valid_percentile_stage_cols = []
        for col_idx, stage_col in enumerate(stage_cols, len(base_cols) + 1):
            percentile_col_name = f'{stage_col}_百分位'
            if percentile_col_name in performance_df.columns:
                cell = ws.cell(row=start_percentile_row, column=col_idx, value=f'{stage_col}百分位')
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                valid_percentile_stage_cols.append(stage_col)
        
        # 写入节点百分位数据（在百分位表头下方）
        for row_idx, (_, row_data) in enumerate(performance_df.iterrows(), start_percentile_row + 1):
            # 基础列位置留空
            for col_idx in range(1, len(base_cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx, value='')
                cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
            
            # 写入阶段百分位数据 - 只写入实际存在的百分位列
            for col_idx, stage_col in enumerate(valid_percentile_stage_cols, len(base_cols) + 1):
                percentile_col = f'{stage_col}_百分位'
                if percentile_col in performance_df.columns:
                    cell_value = row_data.get(percentile_col, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
        
        # 应用条件格式
        self._apply_performance_formatting(ws, performance_df, base_cols, stage_cols, start_percentile_row)
        
        # 优化列宽和行高
        self._optimize_column_widths(ws, base_cols, stage_cols)


    def _apply_performance_formatting(self, ws, performance_df: pd.DataFrame, base_cols: List[str], stage_cols: List[str], start_percentile_row: int) -> None:
        """应用性能分析表的条件格式"""
        # 定义颜色
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # 性能等级着色（耗时数据部分）
        performance_series = performance_df['性能等级']
        for row_idx, perf_value in enumerate(performance_series, 2):
            cell = ws.cell(row=row_idx, column=3)  # 性能等级列
            if perf_value == '优秀':
                cell.fill = green_fill
            elif perf_value == '良好':
                cell.fill = light_green_fill
            elif perf_value == '中等':
                cell.fill = yellow_fill
            elif perf_value == '待优化':
                cell.fill = red_fill
        
        # 瓶颈标记着色（耗时数据部分）- 节点名称和瓶颈标记列使用相同格式
        bottleneck_series = performance_df['瓶颈标记']
        for row_idx, is_bottleneck in enumerate(bottleneck_series, 2):
            if is_bottleneck:
                # 节点名称列（第1列）
                node_cell = ws.cell(row=row_idx, column=1)
                node_cell.fill = red_fill
                
                # 瓶颈标记列（第4列）
                bottleneck_cell = ws.cell(row=row_idx, column=4)
                bottleneck_cell.fill = red_fill
        
        # 综合性能评分着色（耗时数据部分）
        score_series = performance_df['综合性能评分']
        for row_idx, score in enumerate(score_series, 2):
            cell = ws.cell(row=row_idx, column=2)  # 综合性能评分列
            if isinstance(score, (int, float)):
                if score >= 80:
                    cell.fill = green_fill
                elif score >= 60:
                    cell.fill = light_green_fill
                elif score >= 40:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
        
        # 所有阶段总耗时（s）着色（耗时越短，颜色越绿）
        total_duration_series = performance_df['所有间隔内代码执行总耗时（s）']
        if len(total_duration_series) > 0:
            min_duration = total_duration_series.min()
            max_duration = total_duration_series.max()
            
            for row_idx, duration in enumerate(total_duration_series, 2):
                cell = ws.cell(row=row_idx, column=5)  # 所有阶段总耗时列
                
                # 计算相对位置（0到1之间，0表示最快，1表示最慢）
                if max_duration > min_duration:
                    relative_position = (duration - min_duration) / (max_duration - min_duration)
                else:
                    relative_position = 0
                
                # 根据相对位置设置颜色（从绿色到红色）
                if relative_position < 0.25:
                    cell.fill = green_fill
                elif relative_position < 0.5:
                    cell.fill = light_green_fill
                elif relative_position < 0.75:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill




    def _optimize_column_widths(self, ws, base_cols: List[str], stage_cols: List[str]) -> None:
        """优化列宽设置"""
        from openpyxl.utils import get_column_letter
        
        # 设置基础列的固定宽度
        base_col_widths = {
            '节点名称': 15,
            '综合性能评分': 12,
            '性能等级': 10,
            '瓶颈标记': 10,
            '所有间隔内代码执行总耗时（s）': 15,
            '综合耗时排名': 12,
            '结果分析': 25
        }
        
        # 应用基础列宽度
        for col_idx, col_name in enumerate(base_cols, 1):
            if col_name in base_col_widths:
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = base_col_widths[col_name]
        
        # 阶段列自适应宽度
        stage_count = len(stage_cols)
        for offset in range(stage_count):
            col_idx = len(base_cols) + 1 + offset
            max_length = 0
            col_letter = get_column_letter(col_idx)
            
            # 检查表头长度
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value:
                header_length = len(str(header_cell.value))
                max_length = max(max_length, header_length)
            
            # 检查数据长度（包括耗时数据和百分位数据）
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            # 设置列宽，最小12，最大25
            adjusted_width = min(max(max_length + 2, 12), 25)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # 设置行高
        ws.row_dimensions[1].height = 25
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20






class CodeReferenceFinder(LogProcessor):
    """代码引用查找器，用于在代码库中查找相关代码段"""
    
    def __init__(self, config: LogAnalyzerConfig):
        super().__init__(config)
        self.additional_files = [
            '/public/home/user/workspace/labs/hcu_megatron/examples/llama/run_llama2_7B.sh',
            '/public/home/user/workspace/labs/hcu_megatron/examples/llama/train_llama2_7b_1nodes.sh'
        ]
    
    def find_code_references(self, result_file: Path, search_dir: Path) -> None:
        """在代码库中查找相关代码引用"""
        if not result_file.exists():
            raise FileNotFoundError(f"结果文件不存在: {result_file}")
        
        py_files = self._get_python_files(search_dir)
        py_files.extend([Path(f) for f in self.additional_files if Path(f).exists()])
        
        # 根据文件类型选择处理方法
        if result_file.suffix.lower() == '.txt':
            self._process_txt_file(result_file, py_files)
        elif result_file.suffix.lower() in ['.xlsx', '.xls']:
            self._process_excel_file(result_file, py_files)
        else:
            self.logger.warning(f"不支持的文件类型: {result_file.suffix}")
    
    def _get_python_files(self, search_dir: Path) -> List[Path]:
        """获取所有Python文件"""
        py_files = []
        
        if search_dir.exists() and search_dir.is_dir():
            try:
                py_files = list(search_dir.rglob('*.py'))
                # 同时查找shell脚本文件
                py_files.extend(list(search_dir.rglob('*.sh')))
            except Exception as e:
                self.logger.error(f"访问目录 {search_dir} 时出错: {e}")
        else:
            self.logger.warning(f"目录 {search_dir} 不存在或不是目录")
        
        return py_files
    
    def _process_txt_file(self, result_file: Path, py_files: List[Path]) -> None:
        """处理TXT文件"""
        for py_file in py_files:
            self.logger.info(f"正在查找代码引用: {py_file}")
            self._process_single_txt_file(result_file, py_file)
    
    def _process_excel_file(self, result_file: Path, py_files: List[Path]) -> None:
        """处理Excel文件"""
        # 加载Excel工作簿
        wb = load_workbook(result_file)
        modified = False
        
        for py_file in py_files:
            self.logger.info(f"正在查找代码引用: {py_file}")
            if self._process_single_excel_file(wb, py_file):
                modified = True
        
        # 保存修改后的Excel文件
        if modified:
            wb.save(result_file)
            self.logger.info(f"已更新Excel文件: {result_file}")
    
    def _process_single_txt_file(self, result_file: Path, py_file: Path) -> None:
        """处理单个Python文件对TXT文件的影响"""
        try:
            with py_file.open('r', encoding='utf-8') as f:
                py_lines = f.readlines()
            
            with result_file.open('r', encoding='utf-8') as f:
                original_lines = f.readlines()
            
            new_lines = []
            for line in original_lines:
                if '->' in line:
                    processed_line = self._process_reference_line(line, py_file, py_lines)
                    new_lines.append(processed_line)
                else:
                    new_lines.append(line)
            
            # 写回结果文件
            with result_file.open('w', encoding='utf-8') as f:
                f.writelines(new_lines)
                
        except Exception as e:
            self.logger.warning(f"处理文件 {py_file} 时出错: {e}")
    
    def _process_single_excel_file(self, wb, py_file: Path) -> bool:
        """处理单个Python文件对Excel文件的影响"""
        try:
            with py_file.open('r', encoding='utf-8') as f:
                py_lines = f.readlines()
            
            modified = False
            
            # 遍历所有工作表
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                if self._process_worksheet(ws, py_file, py_lines):
                    modified = True
            
            return modified
                
        except Exception as e:
            self.logger.warning(f"处理文件 {py_file} 时出错: {e}")
            return False
    
    def _process_worksheet(self, ws, py_file: Path, py_lines: List[str]) -> bool:
        """处理工作表中的单元格"""
        modified = False
        alignment_style = Alignment(
            horizontal='center',
            vertical='top',
            wrap_text=True
        )
        
        # 遍历所有行和列
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and '->' in str(cell.value):
                    processed_value = self._process_reference_line(str(cell.value), py_file, py_lines)
                    if processed_value != cell.value:
                        cell.value = processed_value
                        # 设置自动换行格式
                        cell.alignment = alignment_style
                        modified = True
        
        return modified
    
    def _process_reference_line(self, text: str, py_file: Path, py_lines: List[str]) -> str:
        """处理引用行"""
        parts = text.split('->')
        if len(parts) < 2:
            return text
        
        _prev = parts[0].strip()
        _next = parts[1].strip()
        
        # 清理可能的额外字符
        _prev = _prev.replace('【', '').replace('】', '').strip()
        _next = _next.replace('【', '').replace('】', '').strip()
        
        prev_pattern = re.compile(rf'\b{re.escape(_prev)}\b')
        next_pattern = re.compile(rf'\b{re.escape(_next)}\b')
        
        prev_line_num, next_line_num = self._find_pattern_lines(
            py_lines, prev_pattern, next_pattern
        )
        
        if prev_line_num != -1 and next_line_num != -1:
            return self._format_code_reference(_prev, _next, py_file, py_lines, 
                                             prev_line_num, next_line_num)
        elif prev_line_num != -1 or next_line_num != -1:
            return self._format_partial_reference(_prev, _next, py_file, 
                                                prev_line_num, next_line_num)
        else:
            return text
    
    def _find_pattern_lines(self, py_lines: List[str], prev_pattern: re.Pattern, 
                          next_pattern: re.Pattern) -> Tuple[int, int]:
        """查找模式匹配的行号"""
        prev_line_num = -1
        next_line_num = -1
        
        for j, py_line in enumerate(py_lines):
            if prev_pattern.search(py_line):
                prev_line_num = j
            if next_pattern.search(py_line):
                next_line_num = j
                if prev_line_num != -1:
                    break
        
        return prev_line_num, next_line_num
    
    def _format_code_reference(self, prev: str, next: str, py_file: Path, 
                             py_lines: List[str], prev_line: int, next_line: int) -> str:
        """格式化代码引用"""
        wrap_content = []
        wrap_str = f' -----------------------------开始：【{prev}】和【{next}】对应代码----------------------------- \n'
        wrap_str += f'【{prev}】和【{next}】所在文件路径：{py_file}，代码如下\n'
        
        if next_line >= (prev_line + 8):
            # 代码太多，省略显示
            wrap_str += f'代码太多中间省略显示，从【{prev_line+1}】行开始到【{next_line+1}】行结束\n'
            wrap_content.extend(py_lines[prev_line:prev_line + 3])
            wrap_content.append('\t\t\t......\n')
            wrap_content.extend(py_lines[next_line - 3:next_line + 1])
        elif prev_line > next_line:
            # 反向引用
            wrap_str += f'代码从【{prev_line+1}】行开始调用前面【{next_line+1}】行的内容\n'
            wrap_content.append(py_lines[prev_line])
            wrap_content.append('\t\t\t.......\n')
            wrap_content.extend(py_lines[next_line - 2:next_line + 1])
        else:
            # 正常范围
            wrap_content.extend(py_lines[prev_line:next_line + 1])
        
        wrap_str += ''.join(wrap_content)
        return wrap_str
    
    def _format_partial_reference(self, prev: str, next: str, py_file: Path, 
                                prev_line: int, next_line: int) -> str:
        """格式化部分引用"""
        if prev_line != -1:
            return f'【{prev}】所在文件路径：【{py_file}】，第【{prev_line+1}】行->{next}'
        else:
            return f'{prev}->【{next}】所在文件路径：【{py_file}】，第【{next_line+1}】行'



class LogAnalysisPipeline:
    """日志分析管道，协调整个分析流程"""
    
    def __init__(self, config: LogAnalyzerConfig):
        self.config = config
        self.extractor = LogExtractor(config)
        self.parser = LogParser(config)
        self.analyzer = LogAnalyzer(config)
        self.code_finder = CodeReferenceFinder(config)
        self.logger = logging.getLogger(self.__class__.__name__)
        self._temp_files: List[Path] = []
    
    def run(self, input_file: Path, start_slogan: str, search_dir: Path, num_nodes: int) -> Tuple[Path, Path]:
        """运行完整的日志分析流程"""
        self.logger.info("开始日志分析流程")
        
        try:
            # 步骤1: 提取日志
            self.logger.info("步骤1: 提取日志行")
            extracted_file = self.extractor.extract_lines_by_slogan(input_file, start_slogan)
            self._temp_files.append(extracted_file)
            
            # 步骤2: 按主机分割日志 - 简化调用
            self.logger.info("步骤2: 按主机分割日志")
            logs_dir = self.extractor.split_logs_by_host(extracted_file, scale_hint=num_nodes)
            self._temp_files.append(logs_dir)
            
            # 步骤3: 解析所有日志文件
            self.logger.info("步骤3: 解析日志文件")
            parsed_count = self._parse_all_logs(logs_dir)
            self.logger.info(f"成功解析 {parsed_count} 个日志文件")
            
            # 步骤4: 分析日志
            self.logger.info("步骤4: 分析日志")
            result_file_txt, result_file_excel = self.analyzer.analyze_logs_directory(logs_dir)
            
            # 步骤5: 查找代码引用
            self.logger.info("步骤5: 查找代码引用")
            # self.code_finder.find_code_references(result_file_txt, search_dir)
            # self.code_finder.find_code_references(result_file_excel, search_dir)
            
            self.logger.info("日志分析流程完成")
            return result_file_txt, result_file_excel
            
        except Exception as e:
            self.logger.error(f"日志分析流程失败: {e}")
            self._cleanup_temp_files()
            raise

    def _cleanup_temp_files(self):
        """清理临时文件"""
        for temp_path in self._temp_files:
            try:
                if temp_path.is_file() and temp_path.exists():
                    temp_path.unlink()
                    self.logger.debug(f"已删除临时文件: {temp_path}")
                elif temp_path.is_dir() and temp_path.exists():
                    shutil.rmtree(temp_path)
                    self.logger.debug(f"已删除临时目录: {temp_path}")
            except Exception as e:
                self.logger.warning(f"删除临时文件 {temp_path} 失败: {e}")

    def _parse_all_logs(self, logs_dir: Path) -> int:
        """解析所有日志文件，返回成功解析的数量"""
        parsed_count = 0
        log_files = [f for f in logs_dir.glob("*.txt") 
                    if not f.name.endswith(getattr(self.config, 'ends_slogan', ''))]
        
        self.logger.info(f"开始解析 {len(log_files)} 个日志文件")
        
        for i, log_file in enumerate(log_files):
            try:
                self.parser.parse_single_log(log_file)
                parsed_count += 1
                
                # 进度报告
                if (i + 1) % 100 == 0 or (i + 1) == len(log_files):
                    self.logger.info(f"解析进度: {i + 1}/{len(log_files)}")
                    
            except Exception as e:
                self.logger.warning(f"解析文件 {log_file.name} 失败: {e}")
                continue
        
        return parsed_count


def main():
    """主函数"""
    config = LogAnalyzerConfig()
    
    parser = argparse.ArgumentParser(description="处理日志文件并进行分析")
    parser.add_argument('--input_file', type=str, required=True, 
                       help="输入日志文件路径")
    parser.add_argument('--start_slogan', type=str, default='training_timestamp',
                       help="用于识别日志行的起始标语")
    parser.add_argument('--dir', type=str, required=True,
                       default=config.default_search_dir,
                       help="提取时间戳的目录")
    parser.add_argument('--lt', type=str, default=config.left_time,
                       help="左时间分隔符")
    parser.add_argument('--gt', type=str, default=config.right_time,
                       help="右时间分隔符")
    parser.add_argument('--slow_threshold', type=float, default=config.slow_stage_threshold,
                       help="慢阶段阈值，单位：秒")
    parser.add_argument('--filter_threshold', type=float, default=config.filter_threshold_seconds,
                       help="过滤阈值，单位：秒")
    parser.add_argument('--interval_time', type=float, default=config.interval_time,
                       help="间隔-代码执行耗时，单位：秒")
    parser.add_argument('--num_nodes', type=int, required=True, default=config.num_nodes,
                       help="节点数量，根据此数量决定使用何种方法处理日志文件")
    
    args = parser.parse_args()
    
    # 更新配置
    config.left_time = args.lt
    config.right_time = args.gt
    config.slow_stage_threshold = args.slow_threshold
    config.filter_threshold_seconds = args.filter_threshold
    config.interval_time = args.interval_time
    
    # 验证输入文件存在
    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"错误: 输入文件不存在: {args.input_file}")
        return 1
    
    # 验证搜索目录存在
    search_path = Path(args.dir)
    if not search_path.exists():
        print(f"警告: 搜索目录不存在: {args.dir}")
    
    try:
        pipeline = LogAnalysisPipeline(config)
        num_nodes = int(args.num_nodes)
        result_file_txt, result_file_excel = pipeline.run(input_path, args.start_slogan, search_path, num_nodes)
        return 0
    except Exception as e:
        print(f"分析过程中出错: {e}")
        return 1


if __name__ == '__main__':
    exit(main())
